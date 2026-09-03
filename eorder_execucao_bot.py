import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext
import threading
import time
import os
import sys
import re
import glob
import json
import unicodedata
import urllib.error
import urllib.request
from datetime import datetime, timedelta, timezone

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException

import openpyxl

EORDER_URL = "https://eorder-ceara.enel.com/geocallcoe/w/index.htm"
NOME_EXPORT = "EXECUCAO"

# ── Publicação automática no painel (Supabase) ────────────────────────
SB_URL = "https://xnkvpxireoosrnrfwcws.supabase.co"
SB_KEY = ("eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6"
          "Inhua3ZweGlyZW9vc3JucmZ3Y3dzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODMxMDA3"
          "NzIsImV4cCI6MjA5ODY3Njc3Mn0.BaCa1dUZAEHhwcqx9Es-U1oXrICk08J14e4mUkieH9g")

# Fechamento diário guardado em snapshots_historico — janela rolante, apaga
# o que passar disso pra não crescer sem limite (ver criar_tabela_historico.sql).
HISTORICO_RETENCAO_DIAS = 30

# Só publica as colunas que o painel web realmente usa — export completo
# tem ~40 colunas e o payload inteiro estoura o timeout de escrita do Supabase.
COLS_TDC = [
    "Numero de Serviço", "Tipo Remessa WIN", " Tipo de Serviço", "Estado TdC",
    "Código Equipe", "Chefe/Responsável de Equipe", "Município", "Endereço Completo",
    "Data Prevista Finalização Trabalhos", "Latitude", "Longitude", "Dica Localização",
    "Rota de Leitura", "Código Cliente", "Nome e Sobrenome Cliente",
]
COLS_EXECUCAO = [
    "Numero de Serviço", "Tipo Remessa WIN", " Tipo de Serviço",
    "Recurso/Equipe", "Município", "Data início Execução", "Data fim Execução", "Código Resultado",
    "Número da Incidencia", "Código Cliente", "Cliente", "Endereço", "Resultado",
    "Nota Codificada",
]

# ── Reincidente: cliente cujo último resultado conhecido foi Improdutivo ──
# Port de RESULTADO_PRODUTIVIDADE/RESULTADO_REGRAS_NOTA/RESULTADO_REGRAS_TEXTO_RUIM/
# _classificarResultado do index.html (~linha 894) -- tem que ficar sincronizado
# com a versão JS se algum código novo for adicionado lá.
RESULTADO_PRODUTIVIDADE = {
    'CJLAE': 'DEVOLVIDO', 'FJL': 'DEVOLVIDO', '0065': 'DEVOLVIDO',
    'RAE11': 'IMPRODUTIVO', 'NRE': 'IMPRODUTIVO', '0032': 'IMPRODUTIVO', 'RAE05': 'IMPRODUTIVO',
    '0096': 'IMPRODUTIVO', 'RAE03': 'IMPRODUTIVO', 'LREC': 'IMPRODUTIVO', 'RAE02': 'IMPRODUTIVO',
    '0097': 'IMPRODUTIVO', '0011': 'PRODUTIVO', '0098': 'IMPRODUTIVO', 'VIUV06': 'IMPRODUTIVO',
    '0040': 'IMPRODUTIVO', '0017': 'IMPRODUTIVO', '0004': 'IMPRODUTIVO', '0015': 'IMPRODUTIVO',
    '0003': 'IMPRODUTIVO', 'RAE01': 'IMPRODUTIVO', '0005': 'IMPRODUTIVO', '0000': 'IMPRODUTIVO',
    '0010': 'IMPRODUTIVO', '0018': 'IMPRODUTIVO', 'RAE06': 'IMPRODUTIVO',
    'VIRV06': 'IMPRODUTIVO', '0007': 'IMPRODUTIVO', '0030': 'IMPRODUTIVO', '0048': 'IMPRODUTIVO',
    'RAE10': 'IMPRODUTIVO', '0033': 'IMPRODUTIVO', 'SDJV06': 'IMPRODUTIVO', 'MGDV5': 'IMPRODUTIVO',
    'SDJE02': 'IMPRODUTIVO', 'EXUE02': 'IMPRODUTIVO', 'MMLE03': 'IMPRODUTIVO', 'MGDV3': 'IMPRODUTIVO',
    'IRME04': 'IMPRODUTIVO', 'SRLV05': 'IMPRODUTIVO', 'DCME02': 'IMPRODUTIVO', 'AMCE03': 'IMPRODUTIVO',
    'STBE03': 'IMPRODUTIVO', 'REA': 'PRODUTIVO', 'SDJV02': 'PRODUTIVO', 'AMCE02': 'PRODUTIVO',
    '0074': 'PRODUTIVO', '0046': 'PRODUTIVO', 'VIRV02': 'PRODUTIVO', 'EXUE01': 'PRODUTIVO',
    'MGDV2': 'PRODUTIVO', 'SEM01': 'PRODUTIVO', 'VIUV02': 'PRODUTIVO', 'VIUV04': 'PRODUTIVO',
    '0020': 'PRODUTIVO', 'MMIE01': 'PRODUTIVO', '0021': 'PRODUTIVO',
    'VIUV05': 'PRODUTIVO', '0001': 'PRODUTIVO',
    'VIRV01': 'PRODUTIVO', 'SRLV02': 'PRODUTIVO', 'VIUV01': 'PRODUTIVO', 'VIRV05': 'PRODUTIVO',
    '0026': 'PRODUTIVO', 'EXRE01': 'PRODUTIVO', '0002': 'PRODUTIVO',
    'SDJE01': 'PRODUTIVO', 'VIRV04': 'PRODUTIVO', 'IRME00': 'PRODUTIVO', '0009': 'PRODUTIVO',
    '0012': 'PRODUTIVO', 'STBE01': 'PRODUTIVO', 'IRME01': 'PRODUTIVO', 'BDTE01': 'PRODUTIVO',
    'ACTV03': 'IMPRODUTIVO', 'MMLE02': 'PRODUTIVO', '0024': 'PRODUTIVO', 'ACTV02': 'PRODUTIVO',
    '0025': 'PRODUTIVO', 'AQME01': 'PRODUTIVO', '0035': 'PRODUTIVO', 'MGDV1': 'PRODUTIVO',
    'SDJV05': 'IMPRODUTIVO', 'CAPE01': 'PRODUTIVO', 'SDJV04': 'PRODUTIVO', 'SDJV03': 'IMPRODUTIVO',
    'PROV01': 'PRODUTIVO', '0052': 'PRODUTIVO', 'SDJV01': 'PRODUTIVO', 'DCME01': 'PRODUTIVO',
    '0068': 'PRODUTIVO', 'IMRE01': 'PRODUTIVO', 'AQME03': 'IMPRODUTIVO', 'SEM03': 'IMPRODUTIVO',
    'IMRE03': 'IMPRODUTIVO', '0019': 'IMPRODUTIVO', 'IMBE02': 'IMPRODUTIVO', 'SEM02': 'IMPRODUTIVO',
    'IRME03': 'IMPRODUTIVO', 'RIME04': 'IMPRODUTIVO', 'ACTV06': 'IMPRODUTIVO', 'SMIE03': 'IMPRODUTIVO',
    '0041': 'IMPRODUTIVO', 'VPCE03': 'IMPRODUTIVO', 'EXRE02': 'IMPRODUTIVO', '0090': 'IMPRODUTIVO',
    '0008': 'IMPRODUTIVO', 'IMBE03': 'IMPRODUTIVO', 'DCME03': 'IMPRODUTIVO', '0043': 'IMPRODUTIVO',
    'IRME02': 'IMPRODUTIVO', 'EXUE03': 'IMPRODUTIVO', 'SRLV04': 'IMPRODUTIVO', 'RNLE01': 'PRODUTIVO',
    'MLDE01': 'PRODUTIVO', '0054': 'PRODUTIVO', '0028': 'PRODUTIVO', 'ACTV01': 'PRODUTIVO',
    '0016': 'PRODUTIVO', 'MMIE02': 'PRODUTIVO', 'IMBE01': 'PRODUTIVO', 'SMIE01': 'PRODUTIVO',
    'ACTE01': 'PRODUTIVO', 'MLDE02': 'IMPRODUTIVO', '0034': 'PRODUTIVO', 'RAE04': 'IMPRODUTIVO',
    '0027': 'IMPRODUTIVO', '0045': 'IMPRODUTIVO', 'LBTVI4': 'IMPRODUTIVO',
    'LBTVI2': 'PRODUTIVO', 'LBTVI5': 'PRODUTIVO', 'LBTVI7': 'IMPRODUTIVO', 'LBTEX4': 'IMPRODUTIVO',
    'LBTEX1': 'PRODUTIVO', '0022': 'PRODUTIVO', 'LBTVI1': 'PRODUTIVO', 'PROV03': 'IMPRODUTIVO',
    'LBTEX2': 'IMPRODUTIVO', 'MGD01': 'PRODUTIVO', 'DBR02': 'PRODUTIVO', 'CMIE01': 'PRODUTIVO',
    'SDJE03': 'IMPRODUTIVO', 'LBTVI3': 'PRODUTIVO', 'MMIE04': 'IMPRODUTIVO', 'LBTEX3': 'IMPRODUTIVO',
    '0029': 'PRODUTIVO', 'DBR03': 'IMPRODUTIVO', 'SRLE01': 'PRODUTIVO', 'SRLV01': 'PRODUTIVO',
    'RNLE02': 'IMPRODUTIVO', 'MGD02': 'IMPRODUTIVO', 'MGD04': 'IMPRODUTIVO', 'VPCE02': 'PRODUTIVO',
    'ACTE02': 'IMPRODUTIVO', 'LNGVI6': 'IMPRODUTIVO', 'TCPE02': 'IMPRODUTIVO', 'TCPE01': 'PRODUTIVO',
    'SRLV03': 'PRODUTIVO', 'ACTV04': 'PRODUTIVO', 'SMIE04': 'IMPRODUTIVO', 'SMA01': 'PRODUTIVO',
    'SMIE02': 'IMPRODUTIVO', 'LNGVI7': 'IMPRODUTIVO', 'BDTE02': 'IMPRODUTIVO', 'ACMV05': 'IMPRODUTIVO',
    'DGD02': 'PRODUTIVO', 'MMIE03': 'IMPRODUTIVO', 'LNGVI2': 'PRODUTIVO', 'RIME02': 'IMPRODUTIVO',
    'TRAN': 'IMPRODUTIVO', '0050': 'PRODUTIVO', '0051': 'PRODUTIVO', '0053': 'PRODUTIVO',
    'MLDE03': 'IMPRODUTIVO', 'RAE07': 'PRODUTIVO', 'ACTE03': 'IMPRODUTIVO', 'ACBV02': 'PRODUTIVO',
    'DGD01': 'PRODUTIVO', 'SEE01': 'PRODUTIVO', 'TMEE01': 'PRODUTIVO', 'IRDE01': 'PRODUTIVO',
    'SML01': 'PRODUTIVO', 'IRCE01': 'PRODUTIVO',
    'APOE01': 'PRODUTIVO', 'VMT01': 'PRODUTIVO', 'VMT02': 'IMPRODUTIVO', 'TMEE02': 'IMPRODUTIVO',
    'TMEE03': 'IMPRODUTIVO', 'DIR02': 'PRODUTIVO', 'TCPE03': 'IMPRODUTIVO', 'SMA03': 'IMPRODUTIVO',
    'CMIE02': 'IMPRODUTIVO', 'MGD05': 'PRODUTIVO', 'ITE01': 'PRODUTIVO', 'DBR01': 'PRODUTIVO',
    'AQME02': 'IMPRODUTIVO', 'IMRE02': 'IMPRODUTIVO', '0070': 'PRODUTIVO', 'ACMV02': 'PRODUTIVO',
    'OUTACM': 'IMPRODUTIVO',
}
RESULTADO_REGRAS_NOTA = {
    'LBTVI6': ['EXTR-NECESSITA EXTENSAO DE REDE'],
    'MMLE04': ['NECESSITA EXTENSAO DE REDE'],
    'ACTV05': ['NECESSITA EXTENSAO DE REDE'],
    'VPCE01': ['VIS EXE- UC NORMAL S/LEVANT CARGA', 'VIST. EXEC.- NECESSITA AFERIÇÃO MEDIDOR'],
    'MMLE01': ['SERVIÇO EXECUTADO'],
}
RESULTADO_REGRAS_TEXTO_RUIM = {'0044': ['IMOVEL FECHADO']}


def _sem_acento(s):
    return ''.join(c for c in unicodedata.normalize('NFD', s) if unicodedata.category(c) != 'Mn').upper()


def _classificar_resultado(r):
    cod = (r.get('Código Resultado') or '').strip().upper()
    regras_nota = RESULTADO_REGRAS_NOTA.get(cod)
    if regras_nota:
        nota = _sem_acento((r.get('Nota Codificada') or '').strip())
        return 'PRODUTIVO' if any(_sem_acento(n) == nota for n in regras_nota) else 'IMPRODUTIVO'
    textos_ruins = RESULTADO_REGRAS_TEXTO_RUIM.get(cod)
    if textos_ruins:
        texto = _sem_acento((r.get('Resultado') or ''))
        return 'IMPRODUTIVO' if any(t in texto for t in textos_ruins) else 'PRODUTIVO'
    return RESULTADO_PRODUTIVIDADE.get(cod)


# Índice local: Código Cliente -> último Improdutivo ainda sem Produtivo
# depois (sem janela de corte -- decisão do Igor: conta desde o início do
# histórico). Mantido incrementalmente pra sempre -- cada rodada só dobra
# em cima do índice já salvo os resultados que acabou de ler, nunca
# recalcula do zero (senão o custo cresceria sem limite com o tempo). Mesma
# lógica do _calcular_os_problematicas do gpm_bot.py (seta em Improdutivo,
# remove em Produtivo), adaptada pra um arquivo único mantido em disco em
# vez de reler N arquivos por dia a cada chamada.
REINCIDENTES_IDX_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "reincidentes_execucao_idx.json")
# Contagem à parte, nunca zera nem quando o cliente resolve com um Produtivo
# (ao contrário do índice acima, que esquece o cliente assim que ele
# resolve) -- é o "desde o início do ano até hoje já deu improdutivo Nx"
# mostrado ao lado do texto Improdutivo no painel.
CONTAGEM_IMPRODUTIVOS_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "contagem_improdutivos_ano_idx.json")


def _carregar_json_idx(caminho):
    try:
        with open(caminho, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def _salvar_json_idx(caminho, dados):
    try:
        with open(caminho, 'w', encoding='utf-8') as f:
            json.dump(dados, f, ensure_ascii=False)
    except Exception:
        pass


def _carregar_indice_reincidentes():
    return _carregar_json_idx(REINCIDENTES_IDX_PATH)


def _salvar_indice_reincidentes(indice):
    _salvar_json_idx(REINCIDENTES_IDX_PATH, indice)


def _parse_data_fim(s):
    try:
        return datetime.strptime((s or '').strip(), "%d/%m/%Y %H:%M")
    except Exception:
        return None


def _atualizar_indice_reincidentes(linhas):
    """Dobra as `linhas` (já lidas pelo robô -- hoje, ou o dia da reconsulta)
    em cima do índice salvo em disco. Ordena por 'Data fim Execução' antes
    de aplicar, pra garantir que o resultado mais recente do lote prevaleça
    mesmo que o export não venha em ordem cronológica. Devolvido ou código
    não mapeado não mexe no índice (mesmo comportamento do gpm_bot.py, sem
    "else").

    Cada rodada do robô relê TODO o "hoje" (não só o que mudou desde a
    última rodada -- ver fazer_busca_execucao), então a mesma conclusão de
    serviço aparece de novo em toda rodada seguinte do mesmo dia. Sem
    tratar isso, cada reprocessamento inflava o "streak" e a contagem do
    ano à toa (contava a mesma visita várias vezes). A trava é comparar a
    'Data fim Execução' exata já registrada -- se for igual, é a mesma
    conclusão sendo relida, pula sem mexer em nada.

    `data_anterior`/`motivo_anterior` guardam a penúltima ocorrência (a de
    ANTES da mais recente) -- só são atualizados quando o dia muda de
    verdade (não a cada rodada do mesmo dia), pra servir de base pro que é
    mostrado como "já teve Improdutivo antes": se o cliente acabou de ser
    marcado Improdutivo HOJE, mostrar de novo a data de hoje ali é
    redundante com o status que a própria linha já mostra -- o que importa
    é a ocorrência anterior a essa."""
    indice = _carregar_indice_reincidentes()
    contagem = _carregar_json_idx(CONTAGEM_IMPRODUTIVOS_PATH)
    ordenadas = sorted(linhas, key=lambda r: _parse_data_fim(r.get('Data fim Execução')) or datetime.min)
    for r in ordenadas:
        cliente = (r.get('Código Cliente') or '').strip()
        if not cliente:
            continue
        p = _classificar_resultado(r)
        nova_data = r.get('Data fim Execução') or ''
        if p == 'IMPRODUTIVO':
            anterior = indice.get(cliente, {})
            if nova_data and anterior.get('data') == nova_data:
                continue  # mesma conclusão já processada numa rodada anterior
            dia_novo = nova_data.split(' ')[0] if nova_data else ''
            dia_ja_registrado = (anterior.get('data') or '').split(' ')[0]
            if anterior and dia_ja_registrado and dia_ja_registrado != dia_novo:
                data_anterior = anterior.get('data')
                motivo_anterior = anterior.get('motivo')
            else:
                data_anterior = anterior.get('data_anterior')
                motivo_anterior = anterior.get('motivo_anterior')
            indice[cliente] = {
                'data': nova_data,
                'motivo': r.get('Resultado') or '',
                'codigo_resultado': r.get('Código Resultado') or '',
                'streak': (anterior.get('streak') or 0) + 1,
                'data_anterior': data_anterior,
                'motivo_anterior': motivo_anterior,
            }
            contagem[cliente] = (contagem.get(cliente) or 0) + 1
        elif p == 'PRODUTIVO':
            indice.pop(cliente, None)
    _salvar_indice_reincidentes(indice)
    _salvar_json_idx(CONTAGEM_IMPRODUTIVOS_PATH, contagem)
    return indice, contagem


def _eh_reincidente_hoje(entry):
    """Um cliente só conta como "reincidente" se já tinha um Improdutivo
    ANTES de hoje sem solução -- um Improdutivo que aconteceu pela
    primeira vez hoje ainda não é reincidência de nada, só vira reincidente
    se continuar sem solução amanhã. Devolve (data, motivo) da ocorrência
    anterior a mostrar, ou None se não conta."""
    if not entry:
        return None
    if entry.get('data_anterior'):
        return entry['data_anterior'], entry.get('motivo_anterior') or ''
    dia_atual = (entry.get('data') or '').split(' ')[0]
    hoje = datetime.now().strftime('%d/%m/%Y')
    if dia_atual and dia_atual != hoje:
        return entry.get('data'), entry.get('motivo') or ''
    return None


# Cache local dos Código Cliente do TdC (Sul) de hoje -- usado só pra saber
# quais clientes filtrar na hora de publicar reincidentes (ver
# _publicar_reincidentes). TdC e Execução rodam em chamadas/threads
# separadas (ver rodar_automatico.py), mas dentro do mesmo ciclo de 30min
# -- salvar aqui (de graça, o TdC já tem `linhas` em mãos) e reler evita um
# GET de ~2,24MB no Supabase a cada rodada (medido: quase 40% da cota
# mensal de egress só nisso). Fica levemente desatualizado entre rodadas
# (o arquivo é do ciclo anterior até o TdC de hoje rodar), mas isso não
# importa pra esse uso -- só precisa estar "aproximadamente certo".
CLIENTES_SUL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "clientes_sul_hoje.json")


def _salvar_clientes_sul_local(clientes):
    _salvar_json_idx(CLIENTES_SUL_PATH, sorted(c for c in clientes if c))


def _carregar_clientes_sul_local():
    return set(_carregar_json_idx(CLIENTES_SUL_PATH) or [])

# ── XPaths ───────────────────────────────────────────────────────────
XP_USER          = "/html/body/table/tbody/tr/td/div/div[2]/div/div/form/div/div[2]/table/tbody/tr[1]/td[2]/input"
XP_PASS          = "/html/body/table/tbody/tr/td/div/div[2]/div/div/form/div/div[2]/table/tbody/tr[2]/td[2]/input"

XP_PLANEJAMENTO  = '//*[@id="TBB_tbm2"]/div[6]'
XP_BUSCA_EXEC    = '//*[@id="TBB_tbm2"]/div[2]'
XP_CENTRO_OP     = '/html/body/div[2]/div/div[2]/div/div[2]/div/div[2]/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/form/table[1]/tbody[1]/tr[2]/td[1]/table/tbody/tr[2]/td[2]/select'
CENTRO_OP_VALOR  = "Cosampa - Sul"
XP_TRES_PONTOS   = '/html/body/div[2]/div/div[2]/div/div[2]/div/div[2]/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/form/table[1]/tbody[1]/tr[2]/td[1]/table/tbody/tr[4]/td[2]/table/tbody/tr/td/div/div[1]/div/div[1]'
XP_CHK_EMERG     = '/html/body/div[2]/div/div[2]/div/div[2]/div/div[2]/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/form/table[1]/tbody[1]/tr[2]/td[1]/table/tbody/tr[4]/td[2]/table/tbody/tr/td/div/div[2]/div/div/div/table/tbody/tr[2]/td[2]'
XP_DATA_EXEC_ROW = '/html/body/div[2]/div/div[2]/div/div[2]/div/div[2]/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/form/table[1]/tbody[1]/tr[2]/td[1]/table/tbody/tr[12]/td[1]'
XP_DATA_INI      = '/html/body/div[2]/div/div[2]/div/div[2]/div/div[2]/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/form/table[1]/tbody[1]/tr[2]/td[1]/table/tbody/tr[12]/td[2]/table/tbody/tr/td[1]/table//input'
XP_DATA_FIM      = '/html/body/div[2]/div/div[2]/div/div[2]/div/div[2]/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/form/table[1]/tbody[1]/tr[2]/td[1]/table/tbody/tr[12]/td[2]/table/tbody/tr/td[3]/table/tbody/tr/td[1]//input'
XP_BTN_BUSCAR    = '/html/body/div[2]/div/div[2]/div/div[2]/div/div[2]/div/div/div[1]/table/tbody/tr/td/div/div[2]/div/form/table[3]/tbody/tr/td/table/tbody/tr/td[3]/button'

XP_BTN_EXPORTAR    = '/html/body/div[2]/div/div[2]/div/div[3]/div[1]/div/div[2]'
XP_CAMPO_NOME_EXP  = '//input[contains(@placeholder, "default")]'
XP_BTN_OK_EXP      = '/html/body/div[2]/div/div[2]/div/div[5]/div/div[2]/div/div/div/div/div[2]/div/form/div[2]/table/tbody/tr/td[1]/button'
XP_FECHAR_MSG_EXP  = '/html/body/div[3]/div[2]/div[1]/div/div[2]'
XP_LISTA_EXPORT    = '/html/body/div[1]/div[2]/table/tbody/tr[3]/td/div/div/table/tbody/tr[1]/td/div[3]/div[2]/div'
XP_TRES_PONTOS_LISTA = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div/form/div/div[1]/div/div[1]'

# ── XPaths — Exportação TdC ─────────────────────────────────────────
NOME_EXPORT_TDC = "CosampaCDU"

XP_LISTA_TDC      = "/html/body/div[1]/div[2]/table/tbody/tr[3]/td/div/div/table/tbody/tr[1]/td/div/div[2]/div[5]"
XP_BUSCA_TDC_MENU = "/html/body/div[1]/div[2]/table/tbody/tr[3]/td/div/div/table/tbody/tr[1]/td/div[2]/div[2]/div[1]"

XP_CENTRO_OP_TDC  = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[3]/tr[6]/td[1]/table/tbody/tr[1]/td/table/tbody/tr[2]/td[2]/select'

XP_TRES_PONTOS_PROC = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[3]/tr[6]/td[1]/table/tbody/tr[5]/td/table/tbody/tr[1]/td/table/tbody/tr/td[3]/div/div[1]/div/div[1]'
XP_CHK_EMERG_TDC    = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[3]/tr[6]/td[1]/table/tbody/tr[5]/td/table/tbody/tr[1]/td/table/tbody/tr/td[3]/div/div[2]/div/div/div/div[2]/div/table/tbody/tr[2]'

XP_TRES_PONTOS_ESTADO = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[3]/tr[6]/td[1]/table/tbody/tr[5]/td/table/tbody/tr[1]/td/table/tbody/tr/td[1]/div/div[1]/div/div[1]'
XP_ESTADO_FINALIZADO = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[3]/tr[6]/td[1]/table/tbody/tr[5]/td/table/tbody/tr[1]/td/table/tbody/tr/td[1]/div/div[2]/div/div/div/div[2]/div/table/tbody/tr[2]'
XP_ESTADO_ANULADO    = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[3]/tr[6]/td[1]/table/tbody/tr[5]/td/table/tbody/tr[1]/td/table/tbody/tr/td[1]/div/div[2]/div/div/div/div[2]/div/table/tbody/tr[3]'
XP_ESTADO_ENCERRADO  = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[3]/tr[6]/td[1]/table/tbody/tr[5]/td/table/tbody/tr[1]/td/table/tbody/tr/td[1]/div/div[2]/div/div/div/div[2]/div/table/tbody/tr[4]'
XP_ESTADO_SUSPENSO   = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[3]/tr[6]/td[1]/table/tbody/tr[5]/td/table/tbody/tr[1]/td/table/tbody/tr/td[1]/div/div[2]/div/div/div/div[2]/div/table/tbody/tr[6]'

XP_DATAS_REF_SPAN = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[9]/tr[1]/td/span'
XP_DATA_LANC_INI  = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[9]/tr[2]/td/table/tbody/tr/td[1]/table/tbody/tr/td/fieldset/table/tbody/tr[1]/td/table/tbody/tr/td[1]//input'
XP_DATA_LANC_FIM  = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[9]/tr[2]/td/table/tbody/tr/td[1]/table/tbody/tr/td/fieldset/table/tbody/tr[3]/td/table/tbody/tr/td[1]//input'

XP_BTN_BUSCAR_TDC = '/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/div[2]/table/tbody/tr/td[2]/button'

XP_TRES_PONTOS_EXPORT_TDC = '/html/body/div[2]/div/div[2]/div/div[3]/div[1]/div/div[1]'
XP_BTN_OK_EXP_TDC         = '/html/body/div[2]/div/div[2]/div/div[16]/div/div[2]/div/div/div/div/div[2]/div/form/div[2]/table/tbody/tr/td[1]/button'
XP_FECHAR_MSG_TDC         = '/html/body/div[7]/div[2]/div[1]/div/div[2]'
XP_LISTA_EXPORT_TDC       = '/html/body/div[1]/div[2]/table/tbody/tr[3]/td/div/div/table/tbody/tr[1]/td/div[2]/div[2]/div[4]'
XP_TRES_PONTOS_LISTA_TDC  = '/html/body/div[2]/div/div[2]/div/div[1]/div/div[1]'


def _xp_linha_arquivo(nome_export):
    return f'//td[starts-with(normalize-space(text()), "{nome_export}_")]'


class EOrderExecucaoBot:
    def __init__(self, log_cb, download_dir, minimizado=False):
        self.log = log_cb
        self.download_dir = download_dir
        self.minimizado = minimizado
        self.driver = None
        self.stop_flag = False
        self._frame_cache = {}

    def _plog(self, msg):
        self.log(msg)

    def _start_driver(self):
        opts = webdriver.ChromeOptions()
        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-extensions")
        opts.add_argument("--disable-infobars")
        opts.page_load_strategy = "eager"
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        opts.add_experimental_option("prefs", {
            "download.default_directory": self.download_dir,
            "download.prompt_for_download": False,
            "credentials_enable_service": False,
            "profile.password_manager_enabled": False,
            "profile.password_manager_leak_detection": False,
        })
        opts.add_argument("--disable-save-password-bubble")
        if sys.platform.startswith("linux"):
            # Chrome recusa a abrir rodando como root sem --no-sandbox; e
            # --disable-dev-shm-usage evita crash em VPS com pouco /dev/shm
            # (RAM compartilhada). Não afeta o comportamento no Windows.
            opts.add_argument("--no-sandbox")
            opts.add_argument("--disable-dev-shm-usage")
        self.driver = webdriver.Chrome(options=opts)
        # No servidor, o Chrome pode ficar mais lento em horas de pico (varios
        # Chromes rodando junto -- monitor de equipes + eOrder) -- 120s
        # (padrão do Selenium) às vezes não é suficiente.
        try:
            self.driver.command_executor.set_timeout(300)
        except Exception:
            pass
        if self.minimizado:
            self.driver.minimize_window()

    def _find(self, xpath, condition=EC.visibility_of_element_located, timeout=20):
        driver = self.driver
        wait_fast = WebDriverWait(driver, min(timeout, 4), poll_frequency=0.2)

        cached = self._frame_cache.get(xpath)
        if cached is not None:
            try:
                driver.switch_to.default_content()
                if cached != "default":
                    driver.switch_to.frame(cached)
                return wait_fast.until(condition((By.XPATH, xpath)))
            except Exception:
                pass

        driver.switch_to.default_content()
        try:
            el = wait_fast.until(condition((By.XPATH, xpath)))
            self._frame_cache[xpath] = "default"
            return el
        except Exception:
            pass

        frames = driver.find_elements(By.TAG_NAME, "iframe")
        for idx, frame in enumerate(frames):
            try:
                driver.switch_to.default_content()
                driver.switch_to.frame(frame)
                el = wait_fast.until(condition((By.XPATH, xpath)))
                self._frame_cache[xpath] = idx
                return el
            except Exception:
                pass

        driver.switch_to.default_content()
        if cached is not None:
            try:
                if cached != "default":
                    driver.switch_to.frame(cached)
                return WebDriverWait(driver, timeout, poll_frequency=0.2).until(
                    condition((By.XPATH, xpath)))
            except Exception:
                pass

        driver.switch_to.default_content()
        raise TimeoutException(f"Elemento não encontrado: {xpath}")

    def _click(self, xpath, timeout=20):
        el = self._find(xpath, EC.element_to_be_clickable, timeout)
        try:
            el.click()
        except Exception:
            self.driver.execute_script("arguments[0].click();", el)
        return el

    def _double_click(self, xpath, timeout=20):
        el = self._find(xpath, EC.element_to_be_clickable, timeout)
        try:
            webdriver.ActionChains(self.driver).double_click(el).perform()
        except Exception:
            self.driver.execute_script(
                "var ev = new MouseEvent('dblclick', {bubbles: true}); arguments[0].dispatchEvent(ev);", el)
        return el

    def _click_por_texto(self, texto):
        """
        Clica no elemento mais específico (mais aninhado) cujo texto exato é
        `texto`, usando JS — evita depender da estrutura exata do DOM de um
        item de menu (ex.: "Atualizar" dentro do menu de 3 pontinhos).
        Retorna True se encontrou e clicou, False caso contrário.
        """
        script = """
            var texto = arguments[0];
            var exatos = document.evaluate(
                "//*[normalize-space(text())=" + JSON.stringify(texto) + "]",
                document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);
            if (exatos.snapshotLength > 0) {
                exatos.snapshotItem(exatos.snapshotLength - 1).click();
                return true;
            }
            var parciais = document.evaluate(
                "//*[contains(normalize-space(text())," + JSON.stringify(texto) + ")]",
                document, null, XPathResult.ORDERED_NODE_SNAPSHOT_TYPE, null);
            if (parciais.snapshotLength === 0) return false;
            var melhor = parciais.snapshotItem(0);
            for (var i = 1; i < parciais.snapshotLength; i++) {
                var el = parciais.snapshotItem(i);
                if (el.textContent.length < melhor.textContent.length) melhor = el;
            }
            melhor.click();
            return true;
        """
        return bool(self.driver.execute_script(script, texto))

    def _type(self, xpath, texto, timeout=20):
        el = self._find(xpath, EC.element_to_be_clickable, timeout)
        el.click(); el.clear(); el.send_keys(str(texto))
        return el

    def _login(self, usuario, senha):
        self._plog("🌐 Abrindo eOrder...")
        self.driver.get(EORDER_URL)
        self._plog("🔑 Fazendo login...")
        campo_user = self._find(XP_USER, EC.visibility_of_element_located, timeout=20)
        campo_user.click(); campo_user.clear(); campo_user.send_keys(usuario)
        campo_pwd = self.driver.find_element(By.XPATH, XP_PASS)
        campo_pwd.click(); campo_pwd.clear(); campo_pwd.send_keys(senha)
        campo_pwd.send_keys(Keys.RETURN)
        self.driver.switch_to.default_content()
        time.sleep(3)
        self._plog("✅ Login efetuado.")

    def _navegar_busca_execucao(self):
        self._plog("📂 Abrindo Planejamento...")
        self._click(XP_PLANEJAMENTO, timeout=20)
        self._plog("📂 Abrindo Busca Execução...")
        self._click(XP_BUSCA_EXEC, timeout=20)

    def _marcar_centro_operativo(self):
        self._plog(f"📍 Marcando Centro Operativo: {CENTRO_OP_VALOR}...")
        el = self._find(XP_CENTRO_OP, EC.presence_of_element_located, timeout=20)
        Select(el).select_by_visible_text(CENTRO_OP_VALOR)

    def _desmarcar_emergencia(self):
        self._plog("⚙️  Abrindo opções de filtro...")
        self._click(XP_TRES_PONTOS, timeout=20)
        self._plog("🚫 Desmarcando 'Atendimento de Emergência'...")
        self._click(XP_CHK_EMERG, timeout=20)

    def _preencher_data(self, data_str):
        self._plog(f"📅 Preenchendo data de execução: {data_str}...")
        self._click(XP_DATA_EXEC_ROW, timeout=20)
        self._type(XP_DATA_INI, data_str, timeout=20)
        self._type(XP_DATA_FIM, data_str, timeout=20)

    def _buscar(self):
        self._plog("🔎 Buscando...")
        self._click(XP_BTN_BUSCAR, timeout=20)

    def _exportar(self):
        self._plog("📤 Exportando resultados...")
        self._click(XP_BTN_EXPORTAR, timeout=20)
        self._plog("   ...clicou no botão de exportar")
        self._exportar_generico(NOME_EXPORT, XP_BTN_OK_EXP, XP_FECHAR_MSG_EXP)

    def _exportar_generico(self, nome_export, xp_btn_ok, xp_fechar_msg):
        """
        Corpo comum de qualquer exportação: digita o nome do arquivo no
        popup "Nome Arquivo", confirma (Enter, com fallback no botão Ok)
        e fecha o popup de confirmação. Quem abre o popup é responsabilidade
        de cada fluxo (botão direto na execução, menu de 3 pontinhos no TdC).
        """
        campo = self._type(XP_CAMPO_NOME_EXP, nome_export, timeout=20)
        self._plog("   ...digitou o nome do arquivo")
        try:
            campo.send_keys(Keys.RETURN)
            self._plog("   ...confirmou com Enter")
        except Exception:
            self._click(xp_btn_ok, timeout=10)
            self._plog("   ...clicou em OK")
        try:
            self._click(xp_fechar_msg, timeout=15)
            self._plog("   ...fechou popup de confirmação")
        except TimeoutException:
            self._plog("⚠️  Popup de confirmação não encontrado, seguindo...")

    @staticmethod
    def _timestamp_de(el):
        texto = (el.text or "").strip()
        m = re.search(r'(\d{8}_\d{6})', texto)
        return m.group(1) if m else ""

    def _tamanho_kb(self, elemento):
        """
        Lê a coluna "Dimensão (KB)" da linha do elemento — o arquivo aparece
        na lista com 0 KB enquanto o servidor ainda está gerando o conteúdo;
        só deve ser baixado depois que esse valor for > 0.
        """
        script = """
            var tr = arguments[0].closest('tr');
            if (!tr) return null;
            var tds = tr.querySelectorAll('td');
            for (var i = 0; i < tds.length; i++) {
                var t = tds[i].textContent.trim();
                if (/^[0-9]+(\\.[0-9]+)?$/.test(t)) { return parseFloat(t); }
            }
            return null;
        """
        return self.driver.execute_script(script, elemento)

    def _elemento_mais_recente(self, xpath, timestamp_limite):
        """
        Entre todos os elementos que casam com xpath (vários arquivos
        EXECUCAO_AAAAMMDD_HHMMSS na lista — outros usuários também exportam
        com esse nome), retorna o de timestamp mais recente, desde que seja
        ESTRITAMENTE maior que `timestamp_limite`.

        `timestamp_limite` é o maior timestamp já presente na lista no
        momento em que abrimos "Lista Exportações" (logo após clicar em
        Exportar) — usamos o relógio do PRÓPRIO SERVIDOR (embutido no nome
        do arquivo) como referência, e não o relógio do PC local, pra não
        sofrer com diferença de horário entre as duas máquinas.
        """
        try:
            self._find(xpath, EC.presence_of_element_located, timeout=3)  # garante frame certo no cache
        except TimeoutException:
            return None  # lista ainda vazia — nenhum arquivo desse nome existe ainda
        elementos = self.driver.find_elements(By.XPATH, xpath)
        novos = [el for el in elementos if self._timestamp_de(el) > timestamp_limite]
        if not novos:
            return None
        return max(novos, key=self._timestamp_de)

    def _baixar_exportacao(self, xp_lista_export, xp_tres_pontos_lista, nome_export,
                            espera_max=600, intervalo=5):
        xp_linha = _xp_linha_arquivo(nome_export)

        self._plog("📋 Abrindo Lista de Exportações...")
        self._click(xp_lista_export, timeout=20)
        try:
            WebDriverWait(self.driver, 10).until(
                EC.invisibility_of_element_located((By.ID, "darkdiv")))
        except TimeoutException:
            pass
        time.sleep(1)

        existentes = self.driver.find_elements(By.XPATH, xp_linha)
        timestamp_limite = max((self._timestamp_de(el) for el in existentes), default="")
        self._plog(f"   (ignorando arquivos com timestamp <= {timestamp_limite or '(nenhum existente)'})")

        self._plog("⏳ Aguardando arquivo aparecer na lista (clicando em Atualizar)...")
        inicio_espera = time.time()
        elemento = None
        while (time.time() - inicio_espera) < espera_max:
            if self.stop_flag:
                self._plog("🛑 Parado.")
                return False
            elemento = self._elemento_mais_recente(xp_linha, timestamp_limite)
            if elemento is not None:
                tamanho = self._tamanho_kb(elemento)
                if tamanho and tamanho > 0:
                    self._plog(f"✅ Arquivo encontrado na lista! ({tamanho:.0f} KB)")
                    break
                self._plog("   ...arquivo apareceu mas ainda está sendo gerado (0 KB), aguardando...")
                elemento = None
            try:
                self._click(xp_tres_pontos_lista, timeout=5)
                time.sleep(1.5)
                if not self._click_por_texto("Atualizar"):
                    # Site pode estar lento pra renderizar o menu — tenta mais
                    # uma vez antes de desistir dessa rodada
                    time.sleep(1.0)
                    if not self._click_por_texto("Atualizar"):
                        self._plog("   ⚠️  Item 'Atualizar' não encontrado no menu.")
            except TimeoutException:
                pass
            time.sleep(intervalo)
            decorrido = time.time() - inicio_espera
            # Se a exportação foi rápida demais, o arquivo já podia estar na
            # lista no instante em que tiramos a "foto" do que já existia
            # (timestamp_limite) — nesse caso nunca vamos achar algo
            # ESTRITAMENTE mais novo. Se já tentamos algumas vezes e o
            # "limite" é muito recente (poucos minutos), assumimos que é o
            # nosso arquivo mesmo.
            if elemento is None and decorrido >= 15 and timestamp_limite:
                try:
                    ts = datetime.strptime(timestamp_limite, "%Y%m%d_%H%M%S")
                    if abs((datetime.now() - ts).total_seconds()) <= 300:
                        candidatos = self.driver.find_elements(By.XPATH, xp_linha)
                        candidato = next(
                            (el for el in candidatos if self._timestamp_de(el) == timestamp_limite), None)
                        if candidato is not None:
                            tamanho = self._tamanho_kb(candidato)
                            if tamanho and tamanho > 0:
                                elemento = candidato
                                self._plog("   (nenhum arquivo mais novo apareceu — assumindo que o mais recente já é o nosso)")
                                self._plog(f"✅ Arquivo encontrado na lista! ({tamanho:.0f} KB)")
                                break
                except ValueError:
                    pass
        else:
            self._plog("❌ Tempo esgotado esperando o arquivo de exportação.")
            return False

        timestamp_baixado = self._timestamp_de(elemento)
        self._plog(f"⬇️  Baixando arquivo mais recente: {elemento.text.strip()}...")
        try:
            webdriver.ActionChains(self.driver).double_click(elemento).perform()
        except Exception:
            self.driver.execute_script(
                "var ev = new MouseEvent('dblclick', {bubbles: true}); arguments[0].dispatchEvent(ev);", elemento)
        return timestamp_baixado or True

    # ── Publicação automática no painel (Supabase) ───────────────────
    def _achar_export_mais_recente(self, prefixo, timestamp_esperado=None, espera_max=60, intervalo=2):
        """
        Espera o arquivo baixado aparecer na pasta de downloads (o navegador
        pode levar alguns segundos pra terminar de gravar em disco) e retorna
        o caminho do arquivo.

        Se `timestamp_esperado` for passado (o timestamp — embutido no nome —
        do arquivo que confirmamos ter clicado pra baixar), só aceitamos um
        arquivo com ESSE timestamp exato no nome, mesmo que exista um arquivo
        mais antigo do mesmo prefixo ainda não limpo na pasta (ex: sobra de
        uma rodada anterior que falhou antes de rodar a limpeza) — sem isso,
        `max(..., key=os.path.getmtime)` podia devolver esse arquivo antigo
        por engano caso ele aparecesse no glob antes do novo terminar de ser
        gravado em disco, publicando dados desatualizados no painel.
        """
        decorrido = 0
        while decorrido < espera_max:
            candidatos = glob.glob(os.path.join(self.download_dir, f"{prefixo}_*.xlsx"))
            candidatos = [c for c in candidatos if not c.endswith(".crdownload")]
            if timestamp_esperado:
                candidatos = [c for c in candidatos if timestamp_esperado in os.path.basename(c)]
            if candidatos:
                return max(candidatos, key=os.path.getmtime)
            time.sleep(intervalo)
            decorrido += intervalo
        return None

    def _limpar_exports_antigos(self, prefixo, manter=1):
        """
        Mantém só o(s) export(s) mais recente(s) de um prefixo na pasta de
        downloads e apaga o resto — sem isso a pasta acumula um arquivo
        novo por hora, todo dia.
        """
        try:
            candidatos = glob.glob(os.path.join(self.download_dir, f"{prefixo}_*.xlsx"))
            candidatos = [c for c in candidatos if not c.endswith(".crdownload")]
            candidatos.sort(key=os.path.getmtime, reverse=True)
            for antigo in candidatos[manter:]:
                try:
                    os.remove(antigo)
                    self._plog(f"🗑️  Removido export antigo: {os.path.basename(antigo)}")
                except OSError as e:
                    self._plog(f"⚠️  Não consegui remover {os.path.basename(antigo)}: {e}")
        except Exception as e:
            self._plog(f"⚠️  Falha ao limpar exports antigos: {e}")

    def _xlsx_para_linhas(self, caminho, colunas=None):
        """
        Lê o xlsx e devolve as linhas como dicts. Se `colunas` for passado,
        filtra só esse subconjunto (usado pro snapshot "leve" que alimenta a
        TV ao vivo); se `colunas` for None, devolve TODAS as colunas do
        arquivo original (usado só no snapshot "completo" pro botão Baixar,
        que não é carregado na tela ao vivo).
        """
        wb = openpyxl.load_workbook(caminho, data_only=True)
        ws = wb.active
        cabecalho = [c.value for c in ws[1]]
        linhas = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            registro = {}
            for h, v in zip(cabecalho, row):
                if h is None or (colunas is not None and h not in colunas):
                    continue
                registro[h] = v
            linhas.append(registro)
        return linhas

    def _post_supabase(self, url, corpo, tentativas=3, espera_s=3):
        """POST com retry -- os erros de SSL/timeout do Supabase que aparecem
        de vez em quando no automatico.log (EOF occurred in violation of
        protocol, canceling statement due to statement timeout) costumam ser
        engasgo passageiro de rede, não falha real. Tenta de novo antes de
        desistir. HTTPError (4xx/5xx com resposta) não tem retry -- é erro do
        próprio Supabase recusando o payload, tentar de novo não muda nada."""
        ultimo_erro = None
        for tentativa in range(1, tentativas + 1):
            req = urllib.request.Request(
                url,
                data=corpo,
                method="POST",
                headers={
                    "apikey": SB_KEY,
                    "Authorization": "Bearer " + SB_KEY,
                    "Content-Type": "application/json",
                    "Prefer": "resolution=merge-duplicates",
                },
            )
            try:
                with urllib.request.urlopen(req) as resp:
                    return resp.status
            except urllib.error.HTTPError:
                raise
            except Exception as e:
                ultimo_erro = e
                if tentativa < tentativas:
                    time.sleep(espera_s)
        raise ultimo_erro

    def _publicar_nuvem(self, regiao, prefixo_arquivo, colunas, timestamp_esperado=None,
                         publicar_ao_vivo=True, data_alvo=None):
        """Publica o snapshot "leve" (só as colunas curadas) que alimenta a
        tela ao vivo, e devolve o caminho do xlsx usado (ou None se falhou)
        pra quem quiser publicar também a versão completa a partir do
        mesmo arquivo, sem procurar de novo.

        `publicar_ao_vivo=False` pula o POST em /snapshots (a tela "ao vivo")
        e só atualiza o snapshots_historico -- usado pra reconsultar um dia
        anterior sem sobrescrever a tela ao vivo com dado velho. `data_alvo`
        é a data (aaaa-mm-dd) sendo fechada no histórico; default é hoje.

        A publicação "ao vivo" e o histórico são tentados independentemente
        -- se o POST da tela ao vivo falhar, o histórico ainda é salvo (senão
        um engasgo de rede no meio do dia derruba os dois de uma vez)."""
        caminho = self._achar_export_mais_recente(prefixo_arquivo, timestamp_esperado=timestamp_esperado)
        if not caminho:
            self._plog(f"⚠️  Não encontrei o arquivo {prefixo_arquivo}*.xlsx pra publicar.")
            return None
        linhas = self._xlsx_para_linhas(caminho, colunas)
        if regiao == "Sul":
            # TdC e Execução rodam em threads/chamadas separadas -- salva os
            # clientes de hoje aqui, de graça (já está com `linhas` em mãos),
            # pra Execução ler depois sem precisar buscar de novo no Supabase
            # (ver _buscar_clientes_sul_hoje: um GET do snapshot Sul inteiro
            # tinha medido 2,24MB -- rodando em toda rodada, uns 2GB/mês só
            # nisso, quase 40% da cota de egress).
            _salvar_clientes_sul_local({(r.get("Código Cliente") or "").strip() for r in linhas if r.get("Código Cliente")})
        # Atualiza o índice de reincidentes com o que acabou de ser lido --
        # roda mesmo quando publicar_ao_vivo=False (reconsulta de ontem
        # também traz resultado válido pra incorporar), antes do POST ao
        # vivo pra não depender dele.
        reincidentes = contagem = None
        if regiao == "Execucao":
            try:
                reincidentes, contagem = _atualizar_indice_reincidentes(linhas)
                self._plog(f"🔁 Índice de reincidentes: {len(reincidentes)} cliente(s) com improdutivo em aberto")
            except Exception as e:
                self._plog(f"⚠️  Falha ao atualizar índice de reincidentes: {e}")
        if publicar_ao_vivo:
            try:
                corpo = json.dumps({
                    "regiao": regiao,
                    "dados": linhas,
                    "atualizado_em": datetime.now(timezone.utc).isoformat(),
                }, default=str).encode("utf-8")
                status = self._post_supabase(SB_URL + "/rest/v1/snapshots", corpo)
                self._plog(f"☁ Painel atualizado ({regiao}): {len(linhas)} registros — status {status}")
            except urllib.error.HTTPError as e:
                self._plog(f"⚠️  Falha ao publicar no painel ({regiao}): {e.code} {e.read().decode('utf-8', 'replace')[:200]}")
            except Exception as e:
                self._plog(f"⚠️  Falha ao publicar no painel ({regiao}): {e}")
            if reincidentes is not None:
                # O índice inteiro (todo cliente com Improdutivo em aberto desde
                # sempre) já passa de 800KB e só cresce -- mandar ele inteiro em
                # toda atualização da tela ao vivo é o mesmo tipo de desperdício
                # que já causou incidente de egress no monitor-desul (ver
                # CLAUDE.md). A tela só usa a parte de clientes que aparecem no
                # Planejado de hoje, então filtra pra só isso antes de publicar.
                clientes_hoje = {(r.get("Código Cliente") or "").strip() for r in linhas if r.get("Código Cliente")}
                clientes_hoje |= _carregar_clientes_sul_local()
                clientes_hoje.discard("")
                self._publicar_reincidentes(regiao, reincidentes, contagem or {}, clientes_hoje)
        self._publicar_historico(regiao, linhas, data_alvo=data_alvo)
        return caminho

    def _publicar_reincidentes(self, regiao, reincidentes, contagem, clientes_relevantes):
        """Publica, como uma linha separada (regiao + "_reincidentes",
        reaproveitando a coluna "dados" que a tabela snapshots já tem --
        colunas fixas, não dá pra acrescentar chave nova no snapshot normal),
        pra cada cliente relevante hoje (ver chamador):
        - se ele conta como reincidente (ver _eh_reincidente_hoje: só entra
          se já tinha Improdutivo ANTES de hoje sem solução -- um Improdutivo
          que aconteceu pela primeira vez hoje não conta ainda), a data e o
          motivo da ocorrência ANTERIOR (não a de hoje, que a própria linha
          do serviço já mostra);
        - `total_ano`: quantas vezes esse cliente já deu Improdutivo desde o
          início do histórico (não zera nem quando ele resolve)."""
        try:
            linhas = []
            for cod in clientes_relevantes:
                anterior = _eh_reincidente_hoje(reincidentes.get(cod))
                total_ano = contagem.get(cod) or 0
                if not anterior and not total_ano:
                    continue
                linha = {"codigo_cliente": cod, "total_ano": total_ano}
                if anterior:
                    linha["data"] = anterior[0]
                    linha["motivo"] = anterior[1]
                linhas.append(linha)
            corpo = json.dumps({
                "regiao": regiao + "_reincidentes",
                "dados": linhas,
                "atualizado_em": datetime.now(timezone.utc).isoformat(),
            }, default=str).encode("utf-8")
            status = self._post_supabase(SB_URL + "/rest/v1/snapshots", corpo)
            self._plog(f"☁ Reincidentes publicados ({regiao}): {len(linhas)} cliente(s) relevantes hoje — status {status}")
        except urllib.error.HTTPError as e:
            self._plog(f"⚠️  Falha ao publicar reincidentes ({regiao}): {e.code} {e.read().decode('utf-8', 'replace')[:200]}")
        except Exception as e:
            self._plog(f"⚠️  Falha ao publicar reincidentes ({regiao}): {e}")

    def _publicar_nuvem_completo(self, regiao, caminho):
        """Publica TODAS as colunas do arquivo original numa região
        separada (ex.: 'Sul_completo') -- só é buscada quando alguém clica
        em "Baixar" no painel, nunca na tela ao vivo, então não pesa o
        carregamento normal da TV."""
        if not caminho:
            return
        try:
            linhas = self._xlsx_para_linhas(caminho, colunas=None)
            corpo = json.dumps({
                "regiao": regiao + "_completo",
                "dados": linhas,
                "atualizado_em": datetime.now(timezone.utc).isoformat(),
            }, default=str).encode("utf-8")
            status = self._post_supabase(SB_URL + "/rest/v1/snapshots", corpo)
            self._plog(f"☁ Versão completa publicada ({regiao}): {len(linhas)} registros — status {status}")
        except urllib.error.HTTPError as e:
            self._plog(f"⚠️  Falha ao publicar versão completa ({regiao}): {e.code} {e.read().decode('utf-8', 'replace')[:200]}")
        except Exception as e:
            self._plog(f"⚠️  Falha ao publicar versão completa ({regiao}): {e}")

    def _historico_existente(self, regiao, data_alvo):
        """Busca quantos registros já estão salvos pra regiao+data_alvo, pra
        servir de referência à trava de _publicar_historico. Retorna None se
        não achar nada salvo ainda (dia novo, sem baseline pra comparar)."""
        try:
            req = urllib.request.Request(
                f"{SB_URL}/rest/v1/snapshots_historico?regiao=eq.{regiao}&data=eq.{data_alvo}&select=dados",
                headers={"apikey": SB_KEY, "Authorization": "Bearer " + SB_KEY},
            )
            with urllib.request.urlopen(req) as resp:
                linhas = json.loads(resp.read())
            if not linhas:
                return None
            return len(linhas[0].get("dados") or [])
        except Exception:
            return None

    def _publicar_historico(self, regiao, linhas, data_alvo=None):
        """
        Guarda o snapshot atual como "fechamento" do dia em
        snapshots_historico (upsert por regiao+data — a última publicação do
        dia é a que fica valendo). `data_alvo` é o dia sendo fechado (formato
        aaaa-mm-dd); se omitido, assume hoje -- mas quando é uma reconsulta de
        um dia anterior (ver rodar_automatico.py), o chamador passa a data
        certa explicitamente, senão o fechamento de ontem seria salvo com a
        chave de hoje por engano.

        Trava: se o novo export vier com bem menos registros do que já
        estava salvo pra esse dia, não sobrescreve -- só avisa no log. Isso
        existe porque uma rodada isolada pode pegar um export vazio/quebrado
        da Enel (rede engasgou, portal deu erro etc.) e, sem essa trava, ela
        apagaria o dia inteiro que já tinha sido capturado certinho pelas
        rodadas anteriores (aconteceu de verdade em 26/07, virou um "buraco"
        permanente no histórico).

        Só o Sul (TdC/Painel Gerencial) apaga fechamentos com mais de
        HISTORICO_RETENCAO_DIAS dias (janela rolante) — o Execucao acumula
        pra sempre, pra alimentar o Painel Operacional (histórico mensal de
        Corte/Recorte/Religação, causa raiz, clientes recorrentes).
        """
        try:
            data_alvo = data_alvo or datetime.now().date().isoformat()
            existentes = self._historico_existente(regiao, data_alvo)
            if existentes is not None and existentes >= 20 and len(linhas) < existentes * 0.5:
                self._plog(
                    f"⚠️  Histórico ({regiao} {data_alvo}) NÃO atualizado: novo export tem só "
                    f"{len(linhas)} registros contra {existentes} já salvos -- provável export "
                    "incompleto da Enel, mantendo o que já tinha."
                )
                return
            corpo = json.dumps({
                "regiao": regiao,
                "data": data_alvo,
                "dados": linhas,
                "salvo_em": datetime.now(timezone.utc).isoformat(),
            }, default=str).encode("utf-8")
            self._post_supabase(SB_URL + "/rest/v1/snapshots_historico", corpo)
            if regiao != "Sul":
                return
            limite = (datetime.now().date() - timedelta(days=HISTORICO_RETENCAO_DIAS)).isoformat()
            req_del = urllib.request.Request(
                f"{SB_URL}/rest/v1/snapshots_historico?regiao=eq.{regiao}&data=lt.{limite}",
                method="DELETE",
                headers={"apikey": SB_KEY, "Authorization": "Bearer " + SB_KEY},
            )
            urllib.request.urlopen(req_del)
        except Exception as e:
            self._plog(f"⚠️  Falha ao salvar histórico ({regiao}): {e}")

    def executar(self, usuario, senha, data_str, data_str_ontem=None, data_alvo_ontem=None, tentativas=3):
        """`data_str_ontem`/`data_alvo_ontem`: quando passados, depois de
        publicar hoje normalmente, reconsulta esse dia anterior só pra
        completar o histórico (sem mexer na tela ao vivo) -- cobre serviços
        fechados depois da última rodada de ontem, que senão nunca seriam
        capturados (o robô só busca "hoje" em cada rodada).

        `tentativas`: o portal da Enel de vez em quando engasga no login ou
        na navegação (menu não termina de renderizar a tempo, sessão cai) --
        já apareceu várias vezes no log como "Elemento não encontrado" ou
        "session not created". Antes, uma falha dessas perdia o ciclo
        inteiro (só tentava de novo 30min depois); agora fecha o navegador e
        tenta de novo do zero (mesmo padrão já usado em _reconsultar_ontem),
        até `tentativas` vezes, antes de desistir de verdade."""
        for tentativa in range(1, tentativas + 1):
            try:
                self._start_driver()
                self._login(usuario, senha)
                self.fazer_busca_execucao(data_str)
                if data_str_ontem:
                    self._reconsultar_ontem(usuario, senha, data_str_ontem, data_alvo_ontem)
                break
            except Exception as e:
                self._plog(f"❌ Erro (tentativa {tentativa}/{tentativas}): {e}")
                try:
                    if self.driver:
                        self.driver.quit()
                except Exception:
                    pass
                self.driver = None
                if tentativa < tentativas:
                    time.sleep(5)
        self._plog("🏁 Finalizado.")

    def _reconsultar_ontem(self, usuario, senha, data_str_ontem, data_alvo_ontem):
        """Reaproveitar a mesma sessão do navegador que acabou de buscar hoje
        pra fazer a segunda busca (de ontem) vinha travando 100% das vezes --
        "Elemento não encontrado" no Centro Operativo ao tentar reabrir Busca
        Execução pela segunda vez (ver automatico.log: acontece todo santo
        dia desde 29/07, quando essa reconsulta foi criada -- ela nunca
        publicou nada de verdade). Em vez de reusar a sessão, fecha o
        navegador e abre um novo do zero, com login novo -- igual uma rodada
        comum -- pra não herdar nenhum popup/iframe que sobrou do fluxo de
        exportação anterior."""
        try:
            if self.driver:
                self.driver.quit()
        except Exception:
            pass
        self._start_driver()
        self._login(usuario, senha)
        self.fazer_busca_execucao(data_str_ontem, publicar_ao_vivo=False, data_alvo=data_alvo_ontem)

    def fazer_busca_execucao(self, data_str, publicar_ao_vivo=True, data_alvo=None):
        """`publicar_ao_vivo=False` + `data_alvo` (aaaa-mm-dd) servem pra
        reconsultar um dia anterior só pra completar o histórico (ver
        rodar_automatico.py), sem sobrescrever a tela ao vivo com dado velho."""
        self._navegar_busca_execucao()
        self._marcar_centro_operativo()
        self._desmarcar_emergencia()
        self._preencher_data(data_str)
        self._buscar()
        self._exportar()
        ok = self._baixar_exportacao(XP_LISTA_EXPORT, XP_TRES_PONTOS_LISTA, NOME_EXPORT)
        if ok:
            self._plog("✅ Fluxo Busca Execução concluído. Aguardando download finalizar...")
            time.sleep(5)
            self._plog(f"💾 Verifique a pasta de downloads: {self.download_dir}")
            timestamp_esperado = ok if isinstance(ok, str) else None
            caminho = self._publicar_nuvem("Execucao", NOME_EXPORT, COLS_EXECUCAO, timestamp_esperado=timestamp_esperado,
                                            publicar_ao_vivo=publicar_ao_vivo, data_alvo=data_alvo)
            if publicar_ao_vivo:
                self._publicar_nuvem_completo("Execucao", caminho)
            self._limpar_exports_antigos(NOME_EXPORT)
        return ok

    # ── Fluxo TdC ────────────────────────────────────────────────────
    def _navegar_busca_tdcs(self):
        self.driver.switch_to.default_content()
        self._plog("📂 Abrindo Lista TdC...")
        self._click(XP_LISTA_TDC, timeout=20)
        self._plog("📂 Abrindo Busca TdCs...")
        self._click(XP_BUSCA_TDC_MENU, timeout=20)

    def _marcar_centro_operativo_tdc(self):
        self._plog(f"📍 Marcando Centro Operativo: {CENTRO_OP_VALOR}...")
        el = self._find(XP_CENTRO_OP_TDC, EC.presence_of_element_located, timeout=20)
        Select(el).select_by_visible_text(CENTRO_OP_VALOR)

    def _processos_subprocessos(self):
        self._plog("⚙️  Abrindo Processos/Subprocessos...")
        self._click(XP_TRES_PONTOS_PROC, timeout=20)
        self._plog("🚫 Desmarcando 'Atendimento de Emergência'...")
        self._click(XP_CHK_EMERG_TDC, timeout=20)

    def _estado_tdc(self):
        self._plog("⚙️  Abrindo Estado de TdC (marcando todos)...")
        self._click(XP_TRES_PONTOS_ESTADO, timeout=20)
        for nome, xp in [("Finalizado", XP_ESTADO_FINALIZADO),
                          ("Anulado", XP_ESTADO_ANULADO),
                          ("Encerrado", XP_ESTADO_ENCERRADO),
                          ("Suspenso", XP_ESTADO_SUSPENSO)]:
            self._plog(f"🚫 Desmarcando '{nome}'...")
            self._click(xp, timeout=20)

    def _datas_referencia_tdc(self):
        self._plog("📅 Abrindo Datas Referência...")
        self._click(XP_DATAS_REF_SPAN, timeout=20)
        hoje = datetime.now().strftime("%d/%m/%Y")
        ha_59_dias = (datetime.now() - timedelta(days=59)).strftime("%d/%m/%Y")
        self._plog(f"📅 Data Lançamento: {ha_59_dias} até {hoje}...")
        self._type(XP_DATA_LANC_INI, ha_59_dias, timeout=20)
        self._type(XP_DATA_LANC_FIM, hoje, timeout=20)

    def _buscar_tdc(self):
        self._plog("🔎 Buscando...")
        self._click(XP_BTN_BUSCAR_TDC, timeout=20)

    def _dump_texto_visivel(self, contendo=None, limite=30):
        """Debug: lista textos curtos visíveis na tela (opcionalmente filtrando
        por substring, case-insensitive) -- usado quando um clique por texto
        falha, pra descobrir se o rótulo mudou no site."""
        script = """
            var filtro = (arguments[0] || "").toLowerCase();
            var els = document.querySelectorAll('body *');
            var vistos = new Set();
            var out = [];
            for (var i = 0; i < els.length && out.length < arguments[1]; i++) {
                var t = (els[i].textContent || "").trim();
                if (!t || t.length > 60) continue;
                if (els[i].children.length > 0) continue;
                if (filtro && t.toLowerCase().indexOf(filtro) === -1) continue;
                if (vistos.has(t)) continue;
                vistos.add(t);
                out.push(t);
            }
            return out;
        """
        try:
            return self.driver.execute_script(script, contendo, limite)
        except Exception:
            return []

    def _exportar_tdc(self):
        self._plog("📤 Exportando TdCs...")
        achou = False
        for tentativa in range(4):
            if tentativa > 0:
                self._plog(f"   🔁 Menu não abriu — reclicando nos 3 pontinhos (tentativa {tentativa+1}/4)...")
                time.sleep(1)
            # Reclica nos 3 pontinhos a cada tentativa -- se o primeiro
            # clique não abrir o menu de verdade (por exemplo, o XPath
            # absoluto acertar um elemento levemente deslocado quando a
            # lista tem muitas páginas/linhas), só reprocurar o texto sem
            # reclicar fica preso pra sempre, já que o menu nunca abriu.
            self._click(XP_TRES_PONTOS_EXPORT_TDC, timeout=20)
            time.sleep(1.5)
            if self._click_por_texto("Exportar em xls"):
                achou = True
                break
        if not achou:
            self._plog("⚠️  Item 'Exportar em xls' não encontrado no menu.")
            textos = self._dump_texto_visivel("export")
            self._plog(f"🔍 Textos visíveis contendo 'export': {textos}")
            raise TimeoutException("Item 'Exportar em xls' não encontrado no menu do TdC.")
        self._plog("   ...abriu menu de exportação")
        self._exportar_generico(NOME_EXPORT_TDC, XP_BTN_OK_EXP_TDC, XP_FECHAR_MSG_TDC)

    def executar_tdc(self, usuario=None, senha=None, tentativas=3):
        """`tentativas` só se aplica quando `usuario` é passado (dono do
        próprio navegador -- ver executar) -- se o portal engasgar no login
        ou na navegação, fecha e tenta de novo do zero em vez de perder o
        ciclo inteiro. Quando chamado com `usuario=None` (reaproveitando um
        navegador já aberto por fora), não tenta de novo -- não é dono do
        navegador pra poder fechá-lo e recomeçar."""
        if usuario is None:
            try:
                self.fazer_tdc()
            except Exception as e:
                self._plog(f"❌ Erro: {e}")
            finally:
                self._plog("🏁 Finalizado.")
            return
        for tentativa in range(1, tentativas + 1):
            try:
                self._start_driver()
                self._login(usuario, senha)
                self.fazer_tdc()
                break
            except Exception as e:
                self._plog(f"❌ Erro (tentativa {tentativa}/{tentativas}): {e}")
                try:
                    if self.driver:
                        self.driver.quit()
                except Exception:
                    pass
                self.driver = None
                if tentativa < tentativas:
                    time.sleep(5)
        self._plog("🏁 Finalizado.")

    def fazer_tdc(self):
        self._navegar_busca_tdcs()
        self._marcar_centro_operativo_tdc()
        self._processos_subprocessos()
        self._estado_tdc()
        self._datas_referencia_tdc()
        self._buscar_tdc()
        self._exportar_tdc()
        # TdC costuma ter uma base bem maior que a Execução — em horários de
        # pico o eOrder demora mais pra gerar o export, então esperamos mais
        # tempo (20 min) antes de desistir, em vez do padrão de 10 min.
        ok = self._baixar_exportacao(XP_LISTA_EXPORT_TDC, XP_TRES_PONTOS_LISTA_TDC, NOME_EXPORT_TDC, espera_max=1200)
        if ok:
            self._plog("✅ Fluxo TdC concluído. Aguardando download finalizar...")
            time.sleep(5)
            self._plog(f"💾 Verifique a pasta de downloads: {self.download_dir}")
            timestamp_esperado = ok if isinstance(ok, str) else None
            caminho = self._publicar_nuvem("Sul", NOME_EXPORT_TDC, COLS_TDC, timestamp_esperado=timestamp_esperado)
            self._publicar_nuvem_completo("Sul", caminho)
            self._limpar_exports_antigos(NOME_EXPORT_TDC)
        return ok


class App(tk.Tk):
    BG       = "#1e1e2e"
    FG       = "#cdd6f4"
    ENTRY_BG = "#313244"
    ACCENT   = "#89b4fa"
    GREEN    = "#a6e3a1"
    RED      = "#f38ba8"

    def __init__(self):
        super().__init__()
        self.title("🤖 eOrder Bot — Execução + TdC — Cosampa Sul")
        self.geometry("560x720")
        self.resizable(False, False)
        self.configure(bg=self.BG)
        self.bots = []
        self._ui()

    def _ui(self):
        BG, FG, ENTRY_BG, ACCENT, GREEN, RED = (
            self.BG, self.FG, self.ENTRY_BG, self.ACCENT, self.GREEN, self.RED)

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("G.TButton", background=GREEN, foreground="#1e1e2e",
                        font=("Segoe UI", 10, "bold"), padding=6)
        style.configure("R.TButton", background=RED, foreground="#1e1e2e",
                        font=("Segoe UI", 10, "bold"), padding=6)
        style.configure("TRadiobutton", background=BG, foreground=FG, font=("Segoe UI", 9))
        style.map("TRadiobutton", background=[("active", BG)])

        tk.Label(self, text="eOrder Bot — Execução + TdC", font=("Segoe UI", 16, "bold"),
                 bg=BG, fg=ACCENT).pack(pady=(16, 2))
        tk.Label(self, text="Busca Execução (EXECUCAO) + Busca TdCs (CosampaCDU) → Cosampa Sul",
                 font=("Segoe UI", 9), bg=BG, fg="#6c7086").pack(pady=(0, 12))

        frm = tk.Frame(self, bg=BG)
        frm.pack(padx=24, fill="x")

        def lbl(t):
            tk.Label(frm, text=t, bg=BG, fg=FG,
                     font=("Segoe UI", 9, "bold"), anchor="w").pack(fill="x", pady=(8, 2))

        def entry(parent, var, show=None):
            e = tk.Entry(parent, textvariable=var, bg=ENTRY_BG, fg=FG,
                         insertbackground=FG, relief="flat",
                         font=("Segoe UI", 9), show=show or "")
            e.pack(fill="x", ipady=5)
            return e

        lbl("⚙️  Modo de execução")
        self.modo_var = tk.StringVar(value="1")
        modo_frame = tk.Frame(frm, bg=BG)
        modo_frame.pack(fill="x")
        ttk.Radiobutton(modo_frame, text="1 acesso (sequencial: Execução depois TdC)",
                        variable=self.modo_var, value="1",
                        command=self._atualizar_modo).pack(anchor="w")
        ttk.Radiobutton(modo_frame, text="2 acessos (paralelo: cada um abre seu eOrder)",
                        variable=self.modo_var, value="2",
                        command=self._atualizar_modo).pack(anchor="w")

        lbl("👤 Usuário (acesso 1 — Busca Execução)")
        self.usuario1_var = tk.StringVar()
        entry(frm, self.usuario1_var)

        lbl("🔑 Senha (acesso 1)")
        self.senha1_var = tk.StringVar()
        entry(frm, self.senha1_var, show="●")

        self.lbl_user2 = tk.Label(frm, text="👤 Usuário (acesso 2 — Busca TdCs)", bg=BG, fg=FG,
                                   font=("Segoe UI", 9, "bold"), anchor="w")
        self.usuario2_var = tk.StringVar()
        self.entry_user2 = tk.Entry(frm, textvariable=self.usuario2_var, bg=ENTRY_BG, fg=FG,
                                     insertbackground=FG, relief="flat", font=("Segoe UI", 9))

        self.lbl_senha2 = tk.Label(frm, text="🔑 Senha (acesso 2)", bg=BG, fg=FG,
                                    font=("Segoe UI", 9, "bold"), anchor="w")
        self.senha2_var = tk.StringVar()
        self.entry_senha2 = tk.Entry(frm, textvariable=self.senha2_var, bg=ENTRY_BG, fg=FG,
                                      insertbackground=FG, relief="flat", font=("Segoe UI", 9), show="●")

        lbl("📅 Data de execução (dd/mm/aaaa) — só para Busca Execução")
        self.data_var = tk.StringVar()
        entry(frm, self.data_var)

        lbl("📁 Pasta de download")
        row_d = tk.Frame(frm, bg=BG)
        row_d.pack(fill="x")
        self.dir_var = tk.StringVar(value=os.path.join(os.path.expanduser("~"), "Downloads"))
        tk.Entry(row_d, textvariable=self.dir_var, bg=ENTRY_BG, fg=FG,
                 insertbackground=FG, relief="flat",
                 font=("Segoe UI", 9)).pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 6))
        tk.Button(row_d, text="Procurar", bg=ACCENT, fg="#1e1e2e",
                  font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2",
                  command=self._escolher_pasta).pack(side="right", ipadx=10, ipady=4)

        bf = tk.Frame(self, bg=BG)
        bf.pack(pady=12, padx=24, fill="x")
        self.btn_ini = ttk.Button(bf, text="▶  INICIAR", style="G.TButton", command=self._iniciar)
        self.btn_ini.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.btn_par = ttk.Button(bf, text="■  PARAR", style="R.TButton",
                                  command=self._parar, state="disabled")
        self.btn_par.pack(side="left", fill="x", expand=True)

        tk.Label(self, text="📋 Log", bg=BG, fg=FG,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=24, pady=(8, 2))
        self.log_box = scrolledtext.ScrolledText(
            self, height=16, bg=ENTRY_BG, fg=FG,
            font=("Consolas", 9), relief="flat", state="disabled")
        self.log_box.pack(fill="both", expand=True, padx=24, pady=(0, 16))

        self._atualizar_modo()

    def _atualizar_modo(self):
        if self.modo_var.get() == "2":
            self.lbl_user2.pack(fill="x", pady=(8, 2))
            self.entry_user2.pack(fill="x", ipady=5)
            self.lbl_senha2.pack(fill="x", pady=(8, 2))
            self.entry_senha2.pack(fill="x", ipady=5)
        else:
            self.lbl_user2.pack_forget()
            self.entry_user2.pack_forget()
            self.lbl_senha2.pack_forget()
            self.entry_senha2.pack_forget()

    def _escolher_pasta(self):
        from tkinter import filedialog
        p = filedialog.askdirectory(title="Selecione a pasta de download")
        if p:
            self.dir_var.set(p)

    def log(self, msg):
        def _do():
            self.log_box.config(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.after(0, _do)

    def _iniciar(self):
        modo = self.modo_var.get()
        usuario1 = self.usuario1_var.get().strip()
        senha1 = self.senha1_var.get().strip()
        data_str = self.data_var.get().strip()
        download_dir = self.dir_var.get().strip()

        if not usuario1 or not senha1:
            messagebox.showerror("Erro", "Preencha usuário e senha do acesso 1.")
            return
        if not data_str:
            messagebox.showerror("Erro", "Preencha a data de execução (dd/mm/aaaa).")
            return
        if not os.path.isdir(download_dir):
            messagebox.showerror("Erro", "Pasta de download inválida.")
            return

        if modo == "2":
            usuario2 = self.usuario2_var.get().strip()
            senha2 = self.senha2_var.get().strip()
            if not usuario2 or not senha2:
                messagebox.showerror("Erro", "Preencha usuário e senha do acesso 2.")
                return

        self.btn_ini.config(state="disabled")
        self.btn_par.config(state="normal")
        self.bots = []

        if modo == "1":
            self.log(f"🚀 Iniciando (1 acesso, sequencial)... data={data_str}")
            bot = EOrderExecucaoBot(lambda m: self.log(f"[Único] {m}"), download_dir)
            self.bots.append(bot)

            def _run():
                try:
                    bot._start_driver()
                    bot._login(usuario1, senha1)
                    bot.fazer_busca_execucao(data_str)
                    bot.fazer_tdc()
                except Exception as e:
                    bot._plog(f"❌ Erro: {e}")
                finally:
                    bot._plog("🏁 Finalizado.")
                    self.after(0, self._on_fim)

            threading.Thread(target=_run, daemon=True).start()

        else:
            self.log(f"🚀 Iniciando (2 acessos, paralelo)... data={data_str}")
            bot1 = EOrderExecucaoBot(lambda m: self.log(f"[Acesso1-Execução] {m}"), download_dir)
            bot2 = EOrderExecucaoBot(lambda m: self.log(f"[Acesso2-TdC] {m}"), download_dir)
            self.bots = [bot1, bot2]

            done = {"n": 0}

            def _marcar_fim():
                done["n"] += 1
                if done["n"] >= 2:
                    self.after(0, self._on_fim)

            def _run1():
                bot1.executar(usuario1, senha1, data_str)
                _marcar_fim()

            def _run2():
                bot2.executar_tdc(usuario2, senha2)
                _marcar_fim()

            threading.Thread(target=_run1, daemon=True).start()
            threading.Thread(target=_run2, daemon=True).start()

    def _on_fim(self):
        self.btn_ini.config(state="normal")
        self.btn_par.config(state="disabled")

    def _parar(self):
        for bot in self.bots:
            bot.stop_flag = True
        self.log("🛑 Sinal de parada enviado...")
        self.btn_par.config(state="disabled")


if __name__ == "__main__":
    App().mainloop()
