import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import time
import os
import re
import openpyxl

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException

EORDER_URL = "https://eorder-ceara.enel.com/geocallcoe/w/index.htm"

# ── XPaths ───────────────────────────────────────────────────────────
XP_USER     = "/html/body/table/tbody/tr/td/div/div[2]/div/div/form/div/div[2]/table/tbody/tr[1]/td[2]/input"
XP_PASS     = "/html/body/table/tbody/tr/td/div/div[2]/div/div/form/div/div[2]/table/tbody/tr[2]/td[2]/input"
XP_LISTA    = "/html/body/div[1]/div[2]/table/tbody/tr[3]/td/div/div/table/tbody/tr[1]/td/div/div[2]/div[5]"
XP_BUSCA_M  = "/html/body/div[1]/div[2]/table/tbody/tr[3]/td/div/div/table/tbody/tr[1]/td/div[2]/div[2]/div[1]"
XP_CAMPO_NS = "/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table/tbody[3]/tr[6]/td[1]/table/tbody/tr[1]/td/table/tbody/tr[2]/td[8]/table/tbody/tr/td/input"
XP_BTN_BSC  = "/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/div[2]/table/tbody/tr/td[2]/button"
XP_BTN_EDI  = "/html/body/div[2]/div/div[2]/div/div[3]/div[2]/div/div/div/table/tbody/tr[2]/td[24]/button"
XP_OBS      = "/html/body/div[2]/div/div[2]/div/div[1]/div[2]/div[1]/div/form/table[1]/tbody[13]/tr[2]/td[2]/textarea"
XP_FECHAR   = "/html/body/div[1]/div[2]/table/tbody/tr[3]/td/div/div/table/tbody/tr[1]/td/div[3]/div[2]/div[8]"

# ── Colunas Excel ────────────────────────────────────────────────────
COL_OS       = 6   # F — número de serviço (entrada)
COL_RESULT   = 2   # B — telefone extraído  (saída)


# ── Extração de telefone ─────────────────────────────────────────────
DDD_PADRAO = "88"  # região Ceará — usado quando o número vem sem DDD

# palavras que indicam onde o telefone está (busca prioriza o texto depois delas)
_RE_KEYWORD_TEL = re.compile(r'(TEL\w*|FONE\w*|CONTATO|WHATSAPP|CEL\w*)', re.IGNORECASE)

# palavras próximas a números que NÃO são telefone (protocolo, ticket, OS, etc.)
_RE_BLACKLIST = re.compile(r'(TICKET|PROT\w*|PROTOCOLO|\bOS\b|AGP|CEP|CNPJ|CPF)', re.IGNORECASE)

# (?<!\d) / (?!\d) em vez de \b: \b não separa dígito de letra (ambos são \w),
# então não detectava números colados a texto, ex.: "TEL: 88981252509Endereco:"
# (?:\+?55)? cobre números com código do país colado, ex.: "+5588981252509"
_RE_TEL_COMPLETO = re.compile(r'(?<!\d)(?:\+?55)?\(?(\d{2})\)?[\s\-\.]?(\d{4,5})[\s\-\.]?(\d{4})(?!\d)')
# número sem DDD (8 ou 9 dígitos) — recebe o DDD_PADRAO
_RE_TEL_SEM_DDD = re.compile(r'(?<!\d)(\d{4,5})[\s\-\.]?(\d{4})(?!\d)')


def _melhor_match(matches):
    """Entre os números encontrados, prioriza o de 9 dígitos (celular)."""
    for ddd, meio, fim in matches:
        if len(meio) == 5:
            return ddd + meio + fim
    ddd, meio, fim = matches[0]
    return ddd + meio + fim


def extrair_telefone(texto):
    """
    Extrai número de telefone brasileiro de texto livre.
    Retorna apenas os dígitos: ex. 88993253791
    """
    if not texto:
        return None

    # 1) se houver palavra-chave de telefone, busca SÓ depois dela
    #    (evita pegar números de protocolo/ticket que vêm antes, ex.: "PROT 1003376547 ... TEL 993504312")
    kw = _RE_KEYWORD_TEL.search(texto)
    if kw:
        janela = texto[kw.end():]
        matches = _RE_TEL_COMPLETO.findall(janela)
        if matches:
            return _melhor_match(matches)
        matches_sem_ddd = _RE_TEL_SEM_DDD.findall(janela)
        if matches_sem_ddd:
            for meio, fim in matches_sem_ddd:
                if len(meio) == 5:
                    return DDD_PADRAO + meio + fim
            meio, fim = matches_sem_ddd[0]
            return DDD_PADRAO + meio + fim

    # 2) sem palavra-chave (ou nada encontrado depois dela): varre o texto todo,
    #    mas descarta números colados a palavras que não são telefone
    proibidas_fim = [b.end() for b in _RE_BLACKLIST.finditer(texto)]
    candidatos = []
    for m in _RE_TEL_COMPLETO.finditer(texto):
        perto_de_proibida = any(abs(m.start() - pe) < 12 for pe in proibidas_fim)
        if not perto_de_proibida:
            candidatos.append(m.groups())
    if not candidatos:
        return None
    for ddd, meio, fim in candidatos:
        if len(meio) == 5:
            return ddd + meio + fim
    # nenhum celular (9 dígitos) entre os candidatos: pega o último (telefone tende a vir no fim)
    ddd, meio, fim = candidatos[-1]
    return ddd + meio + fim


# ── Bot ──────────────────────────────────────────────────────────────
class EOrderBot:
    def __init__(self, log_cb, worker_id=""):
        self.log        = log_cb
        self.worker_id  = worker_id
        self.driver     = None
        self.stop_flag  = False
        self._frame_cache = {}   # xpath -> índice do frame (None = default content)

    def _start_driver(self):
        opts = webdriver.ChromeOptions()
        opts.add_argument("--start-maximized")
        opts.add_argument("--disable-extensions")
        opts.add_argument("--disable-infobars")
        opts.add_argument("--incognito")
        opts.page_load_strategy = "eager"
        opts.add_experimental_option("excludeSwitches", ["enable-automation"])
        opts.add_experimental_option("useAutomationExtension", False)
        self.driver = webdriver.Chrome(options=opts)

    def _find(self, xpath, condition=EC.visibility_of_element_located, timeout=20):
        driver = self.driver
        wait_fast = WebDriverWait(driver, min(timeout, 4), poll_frequency=0.2)

        # 1) tenta primeiro o frame onde esse xpath foi encontrado da última vez
        cached = self._frame_cache.get(xpath)
        if cached is not None:
            try:
                driver.switch_to.default_content()
                if cached != "default":
                    driver.switch_to.frame(cached)
                return wait_fast.until(condition((By.XPATH, xpath)))
            except Exception:
                pass

        # 2) tenta o conteúdo principal
        driver.switch_to.default_content()
        try:
            el = wait_fast.until(condition((By.XPATH, xpath)))
            self._frame_cache[xpath] = "default"
            return el
        except Exception:
            pass

        # 3) escaneia todos os iframes
        frames = driver.find_elements(By.TAG_NAME, "iframe")
        for idx, frame in enumerate(frames):
            try:
                driver.switch_to.default_content()
                driver.switch_to.frame(frame)
                el = wait_fast.until(condition((By.XPATH, xpath)))
                self._frame_cache[xpath] = idx
                return el
            except Exception:
                pass

        # 4) última tentativa com o timeout completo, no frame em cache (se houver)
        driver.switch_to.default_content()
        if cached is not None:
            try:
                if cached != "default":
                    driver.switch_to.frame(cached)
                return WebDriverWait(driver, timeout, poll_frequency=0.2).until(
                    condition((By.XPATH, xpath)))
            except Exception:
                pass

        driver.switch_to.default_content()
        raise TimeoutException(f"Elemento não encontrado: {xpath}")

    def _click(self, xpath, timeout=20):
        el = self._find(xpath, EC.element_to_be_clickable, timeout)
        el.click()
        return el

    def _type(self, xpath, texto, timeout=20):
        el = self._find(xpath, EC.element_to_be_clickable, timeout)
        el.click(); el.clear(); el.send_keys(str(texto))
        return el

    def _get_text(self, xpath, timeout=20):
        el = self._find(xpath, EC.presence_of_element_located, timeout)
        return (el.get_attribute("value") or el.text or "").strip()

    def _plog(self, msg):
        self.log(f"{self.worker_id} {msg}")

    def _login(self, usuario, senha):
        self._plog("🌐 Abrindo eOrder...")
        self.driver.get(EORDER_URL)
        self._plog("🔑 Fazendo login...")
        campo_user = self._find(XP_USER, EC.visibility_of_element_located, timeout=20)
        campo_user.click(); campo_user.clear(); campo_user.send_keys(usuario)
        campo_pwd = self.driver.find_element(By.XPATH, XP_PASS)
        campo_pwd.click(); campo_pwd.clear(); campo_pwd.send_keys(senha)
        campo_pwd.send_keys(Keys.RETURN)
        self.driver.switch_to.default_content()
        try:
            WebDriverWait(self.driver, 20).until(
                EC.presence_of_element_located((By.XPATH, XP_LISTA)))
        except TimeoutException:
            time.sleep(3)
        self._plog("✅ Login efetuado.")

    def _abrir_busca_tdcs(self):
        self.driver.switch_to.default_content()
        self._click(XP_LISTA, timeout=15)
        self._click(XP_BUSCA_M, timeout=15)
        self._find(XP_CAMPO_NS, EC.element_to_be_clickable, timeout=15)
        self._plog("✅ Busca TdCs aberta.")

    def _buscar_os(self, numero_os):
        self._type(XP_CAMPO_NS, numero_os, timeout=10)
        self._click(XP_BTN_BSC, timeout=10)
        self._find(XP_BTN_EDI, EC.element_to_be_clickable, timeout=20)

    def _clicar_editar(self):
        self._click(XP_BTN_EDI, timeout=15)
        self._find(XP_OBS, EC.presence_of_element_located, timeout=25)

    def _processar_os(self, numero_os):
        """
        Executa busca + abrir edição + extrair observação para uma OS.
        Faz 1 nova tentativa automática se pegar 'stale element' (elemento
        recriado pelo JS do eOrder entre o find() e o uso) antes de desistir.
        """
        tentativas = 2
        for tentativa in range(1, tentativas + 1):
            try:
                self._buscar_os(numero_os)
                self._clicar_editar()
                return self._extrair_obs()
            except StaleElementReferenceException:
                if tentativa == tentativas:
                    raise
                self._plog(f"   🔁 Elemento desatualizado, tentando OS {numero_os} de novo...")
                self._frame_cache.clear()
                try:
                    self._fechar()
                except Exception:
                    self._abrir_busca_tdcs()

    def _extrair_obs(self):
        self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
        return self._get_text(XP_OBS, timeout=15)

    def _fechar(self):
        try:
            self._click(XP_FECHAR, timeout=10)
            self._find(XP_CAMPO_NS, EC.element_to_be_clickable, timeout=15)
        except Exception:
            self._plog("   ↩️  Reabrindo busca...")
            self._abrir_busca_tdcs()

    def processar_subset(self, usuario, senha, linhas_data, ws, excel_lock, shared):
        """
        linhas_data : lista de (row_num, numero_os) pré-lida do Excel.
        Resultado gravado na coluna B (COL_RESULT) de cada linha.
        """
        try:
            self._start_driver()
            self._login(usuario, senha)
            self._abrir_busca_tdcs()

            for row_num, numero_os in linhas_data:
                if self.stop_flag:
                    self._plog("🛑 Parado.")
                    break

                self._plog(f"OS: {numero_os}")

                try:
                    obs      = self._processar_os(numero_os)
                    telefone = extrair_telefone(obs)

                    if telefone:
                        resultado = telefone
                        self._plog(f"   📱 {telefone}")
                    else:
                        resultado = "SEM NÚMERO DE TELEFONE NA OBSERVAÇÃO"
                        preview = (obs[:50] + "...") if len(obs) > 50 else (obs or "(vazio)")
                        self._plog(f"   ⚠️  Sem telefone. Obs: {preview}")

                    with excel_lock:
                        ws.cell(row=row_num, column=COL_RESULT).value = resultado

                    self._fechar()

                except Exception as e:
                    self._plog(f"   ⚠️  Erro OS {numero_os}: {e}")
                    with excel_lock:
                        ws.cell(row=row_num, column=COL_RESULT).value = f"ERRO: {str(e)[:120]}"
                    with excel_lock:
                        shared["erros"] += 1
                    # elemento "stale" pode deixar o cache de frames apontando pro lugar errado
                    self._frame_cache.clear()
                    try:
                        self._fechar()
                    except Exception:
                        try:
                            self._abrir_busca_tdcs()
                        except Exception:
                            # recuperação leve falhou — provavelmente ficou um modal travado
                            # na tela; faz reset completo: recarrega e refaz login
                            try:
                                self._plog("   🔄 Recuperando sessão (reload completo)...")
                                self._frame_cache.clear()
                                self.driver.get(EORDER_URL)
                                self._login(usuario, senha)
                                self._abrir_busca_tdcs()
                            except Exception as e2:
                                self._plog(f"   ❌ Falha na recuperação: {e2}")

                with excel_lock:
                    shared["done"] += 1

        except Exception as e:
            self._plog(f"❌ Erro geral: {e}")
        finally:
            with excel_lock:
                shared["workers_done"] += 1
            if self.driver:
                try:
                    self.driver.quit()
                except Exception:
                    pass


# ── Interface ────────────────────────────────────────────────────────
class App(tk.Tk):
    BG       = "#1e1e2e"
    FG       = "#cdd6f4"
    ENTRY_BG = "#313244"
    ACCENT   = "#89b4fa"
    GREEN    = "#a6e3a1"
    RED      = "#f38ba8"

    def __init__(self):
        super().__init__()
        self.title("🤖 eOrder TdC Bot — Multi-Acesso")
        self.geometry("700x660")
        self.resizable(False, False)
        self.configure(bg=self.BG)
        self.excel_path = tk.StringVar()
        self.logins     = []
        self.bots       = []
        self._ui()

    def _ui(self):
        BG, FG, ENTRY_BG, ACCENT, GREEN, RED = (
            self.BG, self.FG, self.ENTRY_BG, self.ACCENT, self.GREEN, self.RED)

        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("G.TButton", background=GREEN, foreground="#1e1e2e",
                        font=("Segoe UI", 10, "bold"), padding=6)
        style.configure("R.TButton", background=RED, foreground="#1e1e2e",
                        font=("Segoe UI", 10, "bold"), padding=6)
        style.configure("TProgressbar", troughcolor=ENTRY_BG, background=ACCENT)

        tk.Label(self, text="eOrder TdC Bot", font=("Segoe UI", 16, "bold"),
                 bg=BG, fg=ACCENT).pack(pady=(16, 2))
        tk.Label(self, text="Extrai telefone das Observações TdC → salva na coluna B (linha 2+)",
                 font=("Segoe UI", 9), bg=BG, fg="#6c7086").pack(pady=(0, 12))

        frm = tk.Frame(self, bg=BG)
        frm.pack(padx=24, fill="x")

        def lbl(t):
            tk.Label(frm, text=t, bg=BG, fg=FG,
                     font=("Segoe UI", 9, "bold"), anchor="w").pack(fill="x", pady=(8, 2))

        lbl("📂 Arquivo Excel (.xlsx)  —  Números de Serviço na coluna F")
        row_e = tk.Frame(frm, bg=BG)
        row_e.pack(fill="x")
        tk.Entry(row_e, textvariable=self.excel_path, bg=ENTRY_BG, fg=FG,
                 insertbackground=FG, relief="flat",
                 font=("Segoe UI", 9)).pack(side="left", fill="x", expand=True, ipady=5, padx=(0, 6))
        tk.Button(row_e, text="Procurar", bg=ACCENT, fg="#1e1e2e",
                  font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2",
                  command=self._escolher).pack(side="right", ipadx=10, ipady=4)

        lbl("🔐 Acessos eOrder  (um Chrome por acesso, em paralelo)")
        logins_outer = tk.Frame(frm, bg=ENTRY_BG)
        logins_outer.pack(fill="x")
        logins_row = tk.Frame(logins_outer, bg=ENTRY_BG)
        logins_row.pack(fill="x")

        self.lb_logins = tk.Listbox(
            logins_row, bg=ENTRY_BG, fg=FG, font=("Segoe UI", 9),
            relief="flat", height=4, selectbackground="#45475a",
            activestyle="none", highlightthickness=0)
        self.lb_logins.pack(side="left", fill="x", expand=True, padx=(4, 0), pady=4)

        btn_col = tk.Frame(logins_row, bg=ENTRY_BG)
        btn_col.pack(side="right", padx=6, pady=4)
        tk.Button(btn_col, text="＋ Adicionar", bg=ACCENT, fg="#1e1e2e",
                  font=("Segoe UI", 9, "bold"), relief="flat", cursor="hand2",
                  command=self._dlg_adicionar).pack(fill="x", ipadx=4, ipady=3)
        tk.Button(btn_col, text="− Remover", bg="#45475a", fg=FG,
                  font=("Segoe UI", 9), relief="flat", cursor="hand2",
                  command=self._remover_acesso).pack(fill="x", ipadx=4, ipady=3, pady=(4, 0))

        bf = tk.Frame(self, bg=BG)
        bf.pack(pady=12, padx=24, fill="x")
        self.btn_ini = ttk.Button(bf, text="▶  INICIAR", style="G.TButton", command=self._iniciar)
        self.btn_ini.pack(side="left", fill="x", expand=True, padx=(0, 8))
        self.btn_par = ttk.Button(bf, text="■  PARAR", style="R.TButton",
                                  command=self._parar, state="disabled")
        self.btn_par.pack(side="left", fill="x", expand=True)

        self.pb = ttk.Progressbar(self, orient="horizontal", mode="determinate",
                                  style="TProgressbar")
        self.pb.pack(fill="x", padx=24, pady=(0, 3))
        self.lbl_pb = tk.Label(self, text="Aguardando início...", bg=BG,
                               fg="#6c7086", font=("Segoe UI", 8))
        self.lbl_pb.pack()

        tk.Label(self, text="📋 Log", bg=BG, fg=FG,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=24, pady=(8, 2))
        self.log_box = scrolledtext.ScrolledText(
            self, height=10, bg=ENTRY_BG, fg=FG,
            font=("Consolas", 9), relief="flat", state="disabled")
        self.log_box.pack(fill="both", padx=24, pady=(0, 16))

    def _dlg_adicionar(self):
        BG, FG, ENTRY_BG, ACCENT = self.BG, self.FG, self.ENTRY_BG, self.ACCENT
        dlg = tk.Toplevel(self)
        dlg.title("Adicionar Acesso")
        dlg.configure(bg=BG)
        dlg.geometry("400x190")
        dlg.resizable(False, False)
        dlg.grab_set()
        dlg.transient(self)

        f = tk.Frame(dlg, bg=BG, padx=20, pady=16)
        f.pack(fill="both", expand=True)

        tk.Label(f, text="Usuário  (ex: ENELINT\\BR00743629)",
                 bg=BG, fg=FG, font=("Segoe UI", 9, "bold")).pack(anchor="w")
        user_var = tk.StringVar()
        tk.Entry(f, textvariable=user_var, bg=ENTRY_BG, fg=FG,
                 insertbackground=FG, relief="flat",
                 font=("Segoe UI", 9)).pack(fill="x", ipady=5)

        tk.Label(f, text="Senha", bg=BG, fg=FG,
                 font=("Segoe UI", 9, "bold")).pack(anchor="w", pady=(10, 0))
        pwd_var = tk.StringVar()
        tk.Entry(f, textvariable=pwd_var, show="●", bg=ENTRY_BG, fg=FG,
                 insertbackground=FG, relief="flat",
                 font=("Segoe UI", 9)).pack(fill="x", ipady=5)

        def confirmar():
            u = user_var.get().strip()
            p = pwd_var.get().strip()
            if not u or not p:
                messagebox.showerror("Erro", "Preencha usuário e senha.", parent=dlg)
                return
            self.logins.append({"usuario": u, "senha": p})
            self.lb_logins.insert("end", f"  Acesso {len(self.logins)}: {u}")
            dlg.destroy()

        tk.Button(f, text="Adicionar", bg=ACCENT, fg="#1e1e2e",
                  font=("Segoe UI", 10, "bold"), relief="flat", cursor="hand2",
                  command=confirmar).pack(pady=(14, 0), ipadx=20, ipady=4)
        dlg.bind("<Return>", lambda e: confirmar())

    def _remover_acesso(self):
        sel = self.lb_logins.curselection()
        if not sel:
            return
        idx = sel[0]
        self.lb_logins.delete(idx)
        self.logins.pop(idx)
        items = list(self.lb_logins.get(0, "end"))
        self.lb_logins.delete(0, "end")
        for i, _ in enumerate(items):
            self.lb_logins.insert("end", f"  Acesso {i+1}: {self.logins[i]['usuario']}")

    def log(self, msg):
        def _do():
            self.log_box.config(state="normal")
            self.log_box.insert("end", msg + "\n")
            self.log_box.see("end")
            self.log_box.config(state="disabled")
        self.after(0, _do)

    def set_progress(self, v):
        def _do():
            self.pb["value"] = v
            self.lbl_pb.config(text=f"Progresso: {v}%")
        self.after(0, _do)

    def _escolher(self):
        p = filedialog.askopenfilename(
            title="Selecione o Excel",
            filetypes=[("Excel", "*.xlsx *.xlsm"), ("Todos", "*.*")])
        if p:
            self.excel_path.set(p)

    def _iniciar(self):
        excel = self.excel_path.get().strip()
        if not excel or not os.path.exists(excel):
            messagebox.showerror("Erro", "Selecione um arquivo Excel válido.")
            return
        if not self.logins:
            messagebox.showerror("Erro", "Adicione pelo menos um acesso.")
            return

        try:
            wb = openpyxl.load_workbook(excel)
            ws = wb.active
            linhas_data = []
            for r in ws.iter_rows(min_row=2):
                if r[COL_OS - 1].value in (None, ""):
                    break
                linhas_data.append((r[COL_OS - 1].row, r[COL_OS - 1].value))
        except Exception as e:
            messagebox.showerror("Erro", f"Não foi possível abrir o Excel:\n{e}")
            return

        if not linhas_data:
            messagebox.showerror("Erro", "Nenhum número encontrado na coluna F.")
            return

        total     = len(linhas_data)
        n_workers = len(self.logins)
        chunks    = [linhas_data[i::n_workers] for i in range(n_workers)]

        excel_lock = threading.Lock()
        shared     = {"done": 0, "erros": 0, "workers_done": 0}

        self.btn_ini.config(state="disabled")
        self.btn_par.config(state="normal")
        self.pb["value"] = 0
        self.log(f"📂 {total} OS | {n_workers} acesso(s) em paralelo")

        self.bots = []
        for i, login in enumerate(self.logins):
            chunk = chunks[i]
            if not chunk:
                with excel_lock:
                    shared["workers_done"] += 1
                continue
            bot = EOrderBot(self.log, worker_id=f"[W{i+1}]")
            self.bots.append(bot)
            threading.Thread(
                target=bot.processar_subset,
                args=(login["usuario"], login["senha"],
                      chunk, ws, excel_lock, shared),
                daemon=True
            ).start()

        self._wb         = wb
        self._excel_path = excel
        self._shared     = shared
        self._total      = total
        self._n_workers  = n_workers
        self._excel_lock = excel_lock
        self._last_save  = time.time()
        self._last_done_saved = 0

        self.after(500, self._watch)

    def _salvar_excel(self):
        try:
            with self._excel_lock:
                self._wb.save(self._excel_path)
        except Exception as e:
            self.log(f"⚠️  Não foi possível salvar agora (arquivo aberto em outro programa?): {e}")

    def _watch(self):
        shared    = self._shared
        total     = self._total
        n_workers = self._n_workers

        with self._excel_lock:
            done         = shared["done"]
            workers_done = shared["workers_done"]
            erros        = shared["erros"]

        self.set_progress(int(done / total * 100) if total else 0)

        # salva incrementalmente: a cada 5s OU a cada 5 novas OS processadas
        agora = time.time()
        if done > self._last_done_saved and (
                agora - self._last_save >= 5 or done - self._last_done_saved >= 5):
            self._salvar_excel()
            self._last_save = agora
            self._last_done_saved = done

        if workers_done >= n_workers:
            try:
                self._wb.save(self._excel_path)
                self.log(f"\n✅ Finalizado! {total - erros}/{total} com sucesso.")
                self.log(f"💾 Salvo: {self._excel_path}")
                if erros:
                    self.log(f"⚠️  {erros} OS com erro — verifique coluna B.")
            except Exception as e:
                self.log(f"❌ Erro ao salvar Excel: {e}")
            self.set_progress(100)
            self.btn_ini.config(state="normal")
            self.btn_par.config(state="disabled")
        else:
            self.after(500, self._watch)

    def _parar(self):
        for bot in self.bots:
            bot.stop_flag = True
        self._salvar_excel()
        self.log("🛑 Sinal de parada enviado para todos os workers... (progresso salvo)")
        self.btn_par.config(state="disabled")


if __name__ == "__main__":
    App().mainloop()