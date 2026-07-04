import openpyxl
import json
import urllib.error
import urllib.request
from datetime import datetime, timezone

SB_URL = 'https://xnkvpxireoosrnrfwcws.supabase.co'
SB_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Inhua3ZweGlyZW9vc3JucmZ3Y3dzIiwicm9sZSI6ImFub24iLCJpYXQiOjE3ODMxMDA3NzIsImV4cCI6MjA5ODY3Njc3Mn0.BaCa1dUZAEHhwcqx9Es-U1oXrICk08J14e4mUkieH9g'


COLS_TDC = [
    'Numero de Serviço', 'Tipo Remessa WIN', ' Tipo de Serviço', 'Estado TdC',
    'Código Equipe', 'Chefe/Responsável de Equipe', 'Município', 'Endereço Completo',
    'Data Prevista Finalização Trabalhos', 'Latitude', 'Longitude', 'Dica Localização',
    'Rota de Leitura', 'Código Cliente', 'Nome e Sobrenome Cliente',
]

COLS_EXECUCAO = [
    'Numero de Serviço', 'Tipo Remessa WIN', ' Tipo de Serviço',
    'Recurso/Equipe', 'Município', 'Data fim Execução',
]


def xlsx_to_rows(path, cols):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for h, v in zip(header, row):
            if h is None or h not in cols:
                continue
            d[h] = v
        rows.append(d)
    return rows


def publicar(regiao, rows):
    body = json.dumps({
        'regiao': regiao,
        'dados': rows,
        'atualizado_em': datetime.now(timezone.utc).isoformat()
    }, default=str).encode('utf-8')
    req = urllib.request.Request(
        SB_URL + '/rest/v1/snapshots',
        data=body,
        method='POST',
        headers={
            'apikey': SB_KEY,
            'Authorization': 'Bearer ' + SB_KEY,
            'Content-Type': 'application/json',
            'Prefer': 'resolution=merge-duplicates',
        },
    )
    print(regiao, '-> payload', round(len(body) / 1024 / 1024, 2), 'MB,', len(rows), 'linhas')
    try:
        with urllib.request.urlopen(req) as resp:
            print(regiao, '-> OK', resp.status)
    except urllib.error.HTTPError as e:
        print(regiao, '-> ERRO', e.code, e.read().decode('utf-8', 'replace'))


if __name__ == '__main__':
    tdc = xlsx_to_rows(r'C:\Users\igor.leite\Downloads\TDCCOSAMPA_20260703_143951.xlsx', COLS_TDC)
    publicar('Sul', tdc)

    exe = xlsx_to_rows(r'C:\Users\igor.leite\Downloads\EXECUCAO_20260626_154836.xlsx', COLS_EXECUCAO)
    publicar('Execucao', exe)
